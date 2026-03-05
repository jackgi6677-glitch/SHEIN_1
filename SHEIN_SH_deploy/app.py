import streamlit as st
import pandas as pd
import json
import os
import re
import copy
import requests
import hmac
import hashlib
import base64
import time
import random
import string
from pathlib import Path
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor

# 飞书等外网请求：遇到 SSL/连接瞬时错误时自动重试（如 SSLEOFError、ConnectionError）
FEISHU_REQUEST_RETRIES = 3
FEISHU_REQUEST_RETRY_DELAY = 1.5

def _request_with_retry(method, url, retries=FEISHU_REQUEST_RETRIES, **kwargs):
    """对 requests 请求做重试，捕获 SSL/连接类异常。"""
    last_err = None
    for attempt in range(retries):
        try:
            if method.upper() == "GET":
                r = requests.get(url, **kwargs)
            else:
                r = requests.post(url, **kwargs)
            return r
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, OSError) as e:
            last_err = e
            if attempt < retries - 1:
                time.sleep(FEISHU_REQUEST_RETRY_DELAY)
            continue
    raise last_err

# ---------- 初始化配置 ----------
st.set_page_config(
    page_title="SHEIN 自动化提效工具 - 模板映射",
    page_icon="⚙️",
    layout="wide"
)

# 常量定义
CONFIG_DIR = Path("templates")
CONFIG_DIR.mkdir(exist_ok=True)

# 选项常量
OP_IGNORE = "无需处理/忽略"
OP_FIXED_SINGLE = "固定单选"
OP_FIXED_MULTI = "固定多选"
OP_FIXED_TEXT = "固定文本填写"
OP_FEISHU = "飞书动态匹配"
OP_IMAGE = "本地图片转直链"
OP_MANUAL = "运行态手动填写"
PROCESSING_OPTIONS = [OP_FIXED_SINGLE, OP_FIXED_MULTI, OP_FIXED_TEXT, OP_FEISHU, OP_IMAGE, OP_MANUAL, OP_IGNORE]

# 核心动态列关键词：运行态由操作员手动录入，配置态仅做标记
MANUAL_KEYWORDS = ["货号", "卖家SKU", "商品名称", "商品描述", "商品卖点", "默认商品"]

# ---------- 内部术语别名映射（UI 显示用，底层 Key 保持原 Excel 列名）----------
ALIAS_MAP = {
    "货号": "供方货号",
    "卖家SKU": "实际货号",
    "默认商品名称": "标题",
    "默认商品描述": "五点描述",
}

def _display_name(col_name):
    """返回 UI 展示名：如果有别名则显示 '别名 (原名)'，否则原样返回。"""
    for orig, alias in ALIAS_MAP.items():
        if orig in col_name:
            return f"{col_name} ({alias})"
    return col_name

# 图片列名关键词：列名包含其一即视为图片列（即使用户未在第一步选「本地图片转直链」也展示上传并写入链接）
IMAGE_COL_KEYWORDS = ["首图", "主图", "细节图", "方形图", "色块图", "详情图", "main_img", "square_img", "detail_img", "color_block"]


def _split_main_detail_image_cols(image_cols):
    """将图片列拆为「首图+细节图1-10」一组（按顺序对应多图上传）、其余为独立列（方块图/色块图/SKU图等）。返回 (main_detail_cols, other_cols)。"""
    if not image_cols:
        return [], []
    main_detail_cols = []
    for c in image_cols:
        if c in ("首图", "主图"):
            main_detail_cols.append(c)
            break
    detail_nums = sorted([int(re.match(r"^细节图(\d+)$", c).group(1)) for c in image_cols if c and re.match(r"^细节图(\d+)$", c)])
    for n in detail_nums:
        col = f"细节图{n}"
        if col in image_cols:
            main_detail_cols.append(col)
    other_cols = [c for c in image_cols if c not in main_detail_cols]
    return main_detail_cols, other_cols

def _image_col_session_keys(col_name):
    """返回可能存放该图片列上传文件的 session key 列表，便于首图/main_img 等别名互通。"""
    keys = [f"img_{col_name}"]
    c = (col_name or "").strip()
    if c:
        keys.append(f"img_{c}")
    col_clean = c.split("\n")[0].strip() if c else ""
    if col_clean and f"img_{col_clean}" not in keys:
        keys.append(f"img_{col_clean}")
    if col_clean == "main_img" or (col_clean and "首图" in col_clean):
        keys.extend(["img_首图", "img_main_img"])
    return list(dict.fromkeys(keys))

def _get_uploaded_for_image_col(col_name):
    """从 session 中取该图片列的上传文件，尝试列名及别名 key。"""
    for k in _image_col_session_keys(col_name):
        v = st.session_state.get(k)
        if v is not None:
            return v
    return None

def _excel_header_row(ws, prefer_row=2, fallback_row=1):
    """按列读取表头：每列优先取 prefer_row 行，为空则取 fallback_row（兼容首图/main_img 等双行或合并表头）。"""
    cols = []
    for col_idx in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=prefer_row, column=col_idx).value
        if not v or (isinstance(v, str) and not v.strip()):
            v = ws.cell(row=fallback_row, column=col_idx).value
        if v is not None:
            v = str(v).strip()
        else:
            v = ""
        cols.append(v)
    return cols

# ---------- 初始化 Session State ----------
if 'excel_columns' not in st.session_state:
    st.session_state.excel_columns = []
if 'config_mapping' not in st.session_state:
    st.session_state.config_mapping = {}
if 'loaded_template_name' not in st.session_state:
    st.session_state.loaded_template_name = ""
if 'feishu_fields' not in st.session_state:
    st.session_state.feishu_fields = []  # 保存从飞书拉取到的真实字段列表

# 获取已保存的本地模板列表
_CREDENTIAL_FILES = {"feishu_auth.json", "shein_auth.json"}

def get_saved_templates():
    """返回模板列表，排除凭证缓存文件（feishu_auth.json / shein_auth.json）。"""
    return [f.name for f in CONFIG_DIR.glob("*.json") if f.name not in _CREDENTIAL_FILES]

# 删除指定模板（同时删除同名 .xlsx 若有）
def delete_template(filename):
    """删除模板 JSON 及同名的 xlsx（若有）。返回 (成功与否, 错误信息)。"""
    if not filename or not filename.endswith(".json"):
        return False, "无效的模板文件名"
    try:
        json_path = CONFIG_DIR / filename
        if json_path.exists():
            json_path.unlink()
        base = Path(filename).stem
        xlsx_path = CONFIG_DIR / f"{base}.xlsx"
        if xlsx_path.exists():
            xlsx_path.unlink()
        return True, None
    except Exception as e:
        return False, str(e)

# 读取本地模板文件
def load_template(filename):
    filepath = CONFIG_DIR / filename
    if filepath.exists():
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def _parse_excel_columns_from_bytes(excel_bytes):
    """从 Excel 字节解析表头、必填、规则，与第一步逻辑一致。返回 (clean_columns, excel_required_dict, excel_rules_dict)。"""
    from io import BytesIO
    try:
        xlsx = pd.ExcelFile(BytesIO(excel_bytes), engine='openpyxl')
        target_sheet = xlsx.sheet_names[1] if len(xlsx.sheet_names) > 1 else xlsx.sheet_names[0]
        df_meta = pd.read_excel(BytesIO(excel_bytes), sheet_name=target_sheet, header=None, nrows=10, engine='openpyxl')
        raw_columns_row = df_meta.iloc[1]
        raw_required_row = df_meta.iloc[3]
        raw_rules_row = df_meta.iloc[5]
        clean_columns = []
        excel_rules_dict = {}
        excel_required_dict = {}
        for col_name, req_val, rule_val in zip(raw_columns_row, raw_required_row, raw_rules_row):
            col_str = str(col_name).strip()
            if col_str and not col_str.startswith('Unnamed') and col_str != 'nan' \
               and not col_str.startswith('请先选择') and not col_str.startswith('请选择'):
                clean_columns.append(col_str)
                excel_rules_dict[col_str] = str(rule_val).strip()
                excel_required_dict[col_str] = str(req_val).strip() if pd.notna(req_val) else ""
        return clean_columns, excel_required_dict, excel_rules_dict
    except Exception:
        return [], {}, {}


def load_template_with_excel(filename):
    """加载 JSON 模板；若存在同名的 .xlsx 则一并加载并返回 (data, excel_bytes 或 None)。"""
    data = load_template(filename)
    base = Path(filename).stem
    xlsx_path = CONFIG_DIR / f"{base}.xlsx"
    excel_bytes = None
    if xlsx_path.exists():
        try:
            with open(xlsx_path, 'rb') as f:
                excel_bytes = f.read()
        except Exception:
            pass
    return data, excel_bytes

# 飞书 API: 获取 Tenant Access Token
def get_feishu_tenant_token(app_id: str, app_secret: str) -> str:
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    headers = {"Content-Type": "application/json; charset=utf-8"}
    payload = {
        "app_id": app_id,
        "app_secret": app_secret
    }
    
    response = _request_with_retry("POST", url, headers=headers, json=payload, timeout=15)
    response.raise_for_status()
    res_json = response.json()
    
    if res_json.get("code") != 0:
        raise Exception(f"获取 Token 失败 (Code {res_json.get('code')}): {res_json.get('msg')}")
        
    return res_json.get("tenant_access_token")

# 飞书 API: 获取多维表格的所有字段
def fetch_feishu_bitable_fields(token: str, app_token: str, table_id: str):
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8"
    }
    
    # 支持分页，单次最多100个字段
    all_fields = []
    page_token = ""
    has_more = True
    
    while has_more:
        params = {"page_size": 100}
        if page_token:
            params["page_token"] = page_token
            
        response = _request_with_retry("GET", url, headers=headers, params=params, timeout=20)
        response.raise_for_status()
        res_json = response.json()
        
        if res_json.get("code") != 0:
            raise Exception(f"获取表格字段失败 (Code {res_json.get('code')}): {res_json.get('msg')}")
            
        data = res_json.get("data", {})
        items = data.get("items", [])
        
        for item in items:
            all_fields.append(item.get("field_name"))
            
        has_more = data.get("has_more", False)
        page_token = data.get("page_token", "")
        
    return all_fields

# 飞书 API: 搜索多维表格记录（按字段精确匹配）
def search_feishu_record(token: str, app_token: str, table_id: str,
                         field_name: str, search_value: str):
    """
    在飞书多维表格中搜索 field_name == search_value 的记录。
    返回第一条匹配记录的 fields 字典；找不到返回 None。
    """
    # 检索值必须非空，否则飞书会返回 InvalidFilter
    _val = (search_value or "").strip()
    if not _val:
        raise ValueError("检索值不能为空，请填写卖家SKU（或复色时的每个SKU）后再生成。")

    if not (field_name or "").strip():
        raise ValueError("飞书检索列不能为空，请在第二步选择「飞书中对应的检索列」。")

    # 使用「列出记录」接口 + filter 查询参数（飞书文档支持）；避免 POST /records/search 的请求体格式导致 400
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8"
    }
    _fname = (field_name or "").strip()
    # 公式：CurrentValue.[字段名]="值"（值内双引号已转义）
    _val_esc = _val.replace("\\", "\\\\").replace('"', '\\"')
    filter_formula = f'CurrentValue.[{_fname}]="{_val_esc}"'
    params = {"page_size": 20, "filter": filter_formula}

    response = _request_with_retry("GET", url, headers=headers, params=params, timeout=20)
    response.raise_for_status()
    res_json = response.json()

    if res_json.get("code") != 0:
        _code = res_json.get("code")
        _msg = res_json.get("msg", "")
        if _code == 1254018 or _code == 1254818 or "InvalidFilter" in (_msg or ""):
            raise Exception(
                f"搜索记录失败 (Code {_code}): {_msg}\n"
                "请检查：①「飞书中对应的检索列」与多维表格中的字段名完全一致（区分中英文）；② 检索列对应的是「文本」类型。"
            )
        raise Exception(f"搜索记录失败 (Code {_code}): {_msg}")

    items = res_json.get("data", {}).get("items", [])
    if not items:
        return None

    # 返回第一条匹配记录的 fields
    record = items[0]
    fields = record.get("fields", {})

    # 飞书返回的字段值可能是复杂对象（如链接、人员等），统一提取为字符串
    clean_fields = {}
    for k, v in fields.items():
        if isinstance(v, list):
            # 多选/多值：取文本拼接
            parts = []
            for item in v:
                if isinstance(item, dict):
                    parts.append(item.get("text", item.get("name", str(item))))
                else:
                    parts.append(str(item))
            clean_fields[k] = ", ".join(parts)
        elif isinstance(v, dict):
            # 单个对象（如人员、链接）
            clean_fields[k] = v.get("text", v.get("name", str(v)))
        else:
            clean_fields[k] = str(v) if v is not None else ""

    return clean_fields


def _upload_to_smms(file_bytes, filename, _retries=2):
    """sm.ms 免费图床，无需 API Key。失败时最多重试 _retries 次。"""
    for attempt in range(_retries + 1):
        try:
            url = "https://sm.ms/api/v2/upload"
            files = {"smfile": (filename or "image.png", file_bytes)}
            resp = requests.post(url, files=files, timeout=15)
            data = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
            if data.get("success") and isinstance(data.get("data"), dict):
                result = (data["data"].get("url") or data["data"].get("urls", {}).get("url") or "").strip()
                if result:
                    return result
            # 重复上传时 sm.ms 返回 images_repeated，从 images 字段里取已有链接
            if data.get("code") == "image_repeated" and isinstance(data.get("images"), str):
                return data["images"].strip()
        except Exception:
            pass
        if attempt < _retries:
            import time as _time
            _time.sleep(1)
    return ""


def _upload_to_catbox(file_bytes, filename, _retries=2):
    """Catbox 免费图床，无需 API Key，作为 sm.ms 的备用。失败时最多重试 _retries 次。"""
    for attempt in range(_retries + 1):
        try:
            url = "https://catbox.moe/user/api.php"
            files = {"fileToUpload": (filename or "image.png", file_bytes)}
            data = {"reqtype": "fileupload"}
            resp = requests.post(url, data=data, files=files, timeout=20)
            if resp.status_code == 200 and resp.text and resp.text.startswith("http"):
                return resp.text.strip()
        except Exception:
            pass
        if attempt < _retries:
            import time as _time
            _time.sleep(1)
    return ""


def _upload_to_0x0(file_bytes, filename, _retries=2):
    """0x0.st 免费图床，无需 API Key，作为第三备用。失败时最多重试 _retries 次。"""
    for attempt in range(_retries + 1):
        try:
            url = "https://0x0.st"
            files = {"file": (filename or "image.png", file_bytes)}
            resp = requests.post(url, files=files, timeout=20)
            if resp.status_code == 200 and resp.text and resp.text.strip().startswith("http"):
                return resp.text.strip()
        except Exception:
            pass
        if attempt < _retries:
            import time as _time
            _time.sleep(1)
    return ""


class _BytesFile:
    """用于从备份 (bytes, filename) 恢复为可上传的对象，兼容 getvalue/name。"""
    def __init__(self, b, name="image.png"):
        self._b = b
        self._n = name
    def getvalue(self):
        return self._b
    @property
    def name(self):
        return self._n


def _upload_image_cols_to_urls_parallel(image_cols, get_uploaded_fn):
    """并行上传多张图到图床，返回 {列名: 直链}。get_uploaded_fn(col_name) 返回该列的上传文件或 None。"""
    tasks = []
    for col_name in image_cols:
        up = get_uploaded_fn(col_name)
        if not up:
            continue
        try:
            if hasattr(up, "seek"):
                up.seek(0)
            fb = up.getvalue() if hasattr(up, "getvalue") else (up.read() if hasattr(up, "read") else None)
            fn = getattr(up, "name", None) or "image.png"
            if fb:
                tasks.append((col_name, fb, fn))
        except Exception:
            pass
    if not tasks:
        return {}

    def _one(args):
        c, b, f = args
        return (c, upload_image_to_url(None, b, f))

    out = {}
    with ThreadPoolExecutor(max_workers=6) as ex:
        for col_name, url in ex.map(_one, tasks):
            out[col_name] = url or ""
    return out


def upload_image_to_url(uploaded_file, file_bytes=None, filename=None):
    """
    将图片上传到免费图床，返回公网可访问的 HTTPS 直链。
    无需 SHEIN 白名单：先试 sm.ms，失败则试 Catbox。
    可传 UploadedFile，或 (file_bytes, filename) 避免流被读尽后无法再次上传。
    :return: str 直链 URL，失败或未传文件时返回 ""
    """
    if file_bytes is not None and filename is not None:
        pass
    elif uploaded_file is not None:
        try:
            if hasattr(uploaded_file, "seek"):
                uploaded_file.seek(0)
            if hasattr(uploaded_file, "getvalue"):
                file_bytes = uploaded_file.getvalue()
            elif hasattr(uploaded_file, "read"):
                file_bytes = uploaded_file.read()
            else:
                return ""
            filename = getattr(uploaded_file, "name", None) or "image.png"
        except Exception:
            return ""
    else:
        return ""
    if not file_bytes:
        return ""
    fn = filename or "image.png"
    out = _upload_to_smms(file_bytes, fn)
    if not out:
        out = _upload_to_catbox(file_bytes, fn)
    if not out:
        out = _upload_to_0x0(file_bytes, fn)
    return out


# ---------- SHEIN 图片 API：直传本地文件(upload-pic) 与 URL 转链(transform-pic) ----------
# 设为 False 时仅用免费图床(sm.ms/Catbox)生成直链，不走 SHEIN 接口；改为 True 可恢复 SHEIN 直传/转链
USE_SHEIN_IMAGE_API = False

SHEIN_UPLOAD_PIC_PATH = "/open-api/goods/upload-pic"
SHEIN_UPLOAD_PIC_URL = "https://openapi.sheincorp.cn" + SHEIN_UPLOAD_PIC_PATH
SHEIN_TRANSFORM_PIC_PATH = "/open-api/goods/transform-pic"
SHEIN_TRANSFORM_PIC_URL = "https://openapi.sheincorp.cn" + SHEIN_TRANSFORM_PIC_PATH

# 列名关键词 → SHEIN image_type：主图1、细节图2、方形图5、色块图6、详情图7
SHEIN_IMAGE_TYPE_MAP = [
    (["主图", "首图"], 1),
    (["细节图"], 2),
    (["方形图"], 5),
    (["色块图"], 6),
    (["详情图"], 7),
]


def _shein_image_type_for_column(col_name):
    """根据列名推断 SHEIN 图片类型；模板可覆盖为 shein_image_type。"""
    if not col_name:
        return 1
    for keywords, itype in SHEIN_IMAGE_TYPE_MAP:
        if any(kw in col_name for kw in keywords):
            return itype
    return 1  # 默认主图


def _generate_shein_signature(open_key_id, secret_key, path, timestamp, random_key):
    """生成 SHEIN 开放平台 API 签名。value = open_key_id & timestamp & path；key = secret_key + random_key；HMAC-SHA256 → hex → base64；最终 = random_key + base64。"""
    value = f"{open_key_id}&{timestamp}&{path}"
    key = f"{secret_key}{random_key}"
    hmac_result = hmac.new(key.encode("utf-8"), value.encode("utf-8"), hashlib.sha256).digest()
    hex_sig = hmac_result.hex()
    b64_sig = base64.b64encode(hex_sig.encode("utf-8")).decode("utf-8")
    return f"{random_key}{b64_sig}"


def _extract_shein_url_from_response(resp_data):
    """从 SHEIN 接口 JSON 响应中提取图片 URL，兼容多种返回结构。返回 (url_str, error_msg)。"""
    if not resp_data:
        return "", "响应为空"
    url = ""
    inner = resp_data.get("data")
    if isinstance(inner, str) and inner.strip().startswith("http"):
        url = inner.strip()
    elif isinstance(inner, dict):
        url = (
            inner.get("url")
            or inner.get("imageUrl")
            or inner.get("image_url")
            or inner.get("transformedUrl")
            or inner.get("transformed_url")
            or ""
        )
        if isinstance(url, str):
            url = url.strip()
        else:
            url = ""
    if not url and isinstance(resp_data.get("url"), str):
        url = resp_data.get("url", "").strip()
    if url:
        return url, None
    code = resp_data.get("code", resp_data.get("status"))
    msg = resp_data.get("msg", resp_data.get("message", ""))
    return "", f"code={code} msg={msg}" if (code is not None or msg) else "响应中无 url 字段"


def shein_upload_pic(open_key_id, secret_key, file_bytes, filename, image_type, language="zh-cn"):
    """
    调用 SHEIN 图片直传接口，将本地文件直接上传为 SHEIN 直链（无需先传图床）。
    遇 SSL/连接瞬时错误时自动重试。
    :return: (url_str, error_msg)，成功时 error_msg 为 None，失败时 url 为空且 error_msg 为原因。
    """
    if not file_bytes or not (open_key_id and secret_key):
        return "", "参数缺失"
    last_err = None
    for attempt in range(FEISHU_REQUEST_RETRIES):
        try:
            timestamp = str(int(time.time() * 1000))
            random_key = "".join(random.choices(string.ascii_letters + string.digits, k=5))
            signature = _generate_shein_signature(
                open_key_id, secret_key, SHEIN_UPLOAD_PIC_PATH, timestamp, random_key
            )
            headers = {
                "language": language,
                "x-lt-openKeyId": open_key_id,
                "x-lt-signature": signature,
                "x-lt-timestamp": timestamp,
                "Host": "openapi.sheincorp.cn",
            }
            files = {"file": (filename or "image.png", file_bytes)}
            data = {"image_type": int(image_type)}
            resp = requests.post(
                SHEIN_UPLOAD_PIC_URL,
                headers=headers,
                data=data,
                files=files,
                timeout=20,
            )
            if resp.status_code != 200:
                try:
                    body = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
                    return "", _extract_shein_url_from_response(body)[1] or f"HTTP {resp.status_code}"
                except Exception:
                    return "", f"HTTP {resp.status_code}"
            resp_data = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
            return _extract_shein_url_from_response(resp_data)
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, OSError) as e:
            last_err = e
            if attempt < FEISHU_REQUEST_RETRIES - 1:
                time.sleep(FEISHU_REQUEST_RETRY_DELAY)
            continue
        except Exception as e:
            return "", str(e)
    return "", str(last_err) if last_err else "连接失败"


def shein_transform_pic(open_key_id, secret_key, original_url, image_type, language="zh-cn"):
    """
    调用 SHEIN 图片转链接口，将外部图片 URL 转为 SHEIN 可用的直链。
    遇 SSL/连接瞬时错误时自动重试。
    :return: (url_str, error_msg)，成功时 error_msg 为 None，失败时 url 为空且 error_msg 为原因。
    """
    if not original_url or not (open_key_id and secret_key):
        return "", "参数缺失"
    last_err = None
    for attempt in range(FEISHU_REQUEST_RETRIES):
        try:
            timestamp = str(int(time.time() * 1000))
            random_key = "".join(random.choices(string.ascii_letters + string.digits, k=5))
            signature = _generate_shein_signature(
                open_key_id, secret_key, SHEIN_TRANSFORM_PIC_PATH, timestamp, random_key
            )
            headers = {
                "language": language,
                "x-lt-openKeyId": open_key_id,
                "x-lt-signature": signature,
                "x-lt-timestamp": timestamp,
                "Content-Type": "application/json",
                "Host": "openapi.sheincorp.cn",
            }
            payload = {"image_type": int(image_type), "original_url": original_url.strip()}
            resp = requests.post(
                SHEIN_TRANSFORM_PIC_URL,
                headers=headers,
                json=payload,
                timeout=15,
            )
            if resp.status_code != 200:
                try:
                    body = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
                    return "", _extract_shein_url_from_response(body)[1] or f"HTTP {resp.status_code}"
                except Exception:
                    return "", f"HTTP {resp.status_code}"
            data = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
            return _extract_shein_url_from_response(data)
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, OSError) as e:
            last_err = e
            if attempt < FEISHU_REQUEST_RETRIES - 1:
                time.sleep(FEISHU_REQUEST_RETRY_DELAY)
            continue
        except Exception as e:
            return "", str(e)
    return "", str(last_err) if last_err else "连接失败"


# ---------- 侧边栏：飞书 / SHEIN 配置区 ----------
FEISHU_CACHE_FILE = CONFIG_DIR / "feishu_auth.json"
SHEIN_CACHE_FILE = CONFIG_DIR / "shein_auth.json"

def _is_deployed():
    """是否处于「已部署」环境：部署时在后台设 DEPLOYED=1，则不再读写飞书凭证文件，每人仅用自己当次填的 API，互不看到。"""
    v = os.environ.get("DEPLOYED", "").strip().lower()
    return v in ("1", "true", "yes")

def _load_feishu_cache():
    """从独立缓存文件加载飞书凭据；部署环境下不读文件，实现环境隔离（每人只看自己的）。"""
    if _is_deployed():
        return {}
    if FEISHU_CACHE_FILE.exists():
        try:
            with open(FEISHU_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_feishu_cache(data):
    """保存飞书凭据到独立缓存文件；部署环境下不写文件，避免同事的 API 被其他人看到。"""
    if _is_deployed():
        return
    try:
        with open(FEISHU_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _load_shein_cache():
    """从本地文件加载 SHEIN 图片转链凭据；部署环境下不读文件。后续可改为仅从后端 Secrets 读取以隐藏。"""
    if _is_deployed():
        return {}
    if SHEIN_CACHE_FILE.exists():
        try:
            with open(SHEIN_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_shein_cache(data):
    """保存 SHEIN 凭据到本地文件；后续可改为仅用后端配置以隐藏侧栏输入。"""
    if _is_deployed():
        return
    try:
        with open(SHEIN_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _get_feishu_from_secrets():
    """若部署时在后台配置了 Secrets 或环境变量，则返回飞书凭据且不在页面展示，避免同事看到。"""
    # 1) Streamlit Cloud：st.secrets（返回 AttrDict，不是 plain dict，用 hasattr 判断）
    try:
        s = getattr(st, "secrets", None)
        if s is not None:
            # 尝试读取 [feishu] 节
            try:
                feishu = s["feishu"]
                result = {
                    "app_id":     str(feishu["app_id"]),
                    "app_secret": str(feishu["app_secret"]),
                    "app_token":  str(feishu["app_token"]),
                    "table_id":   str(feishu["table_id"]),
                }
                if all(result.values()):
                    return result
            except (KeyError, TypeError):
                pass
            # 兼容平铺写法：feishu_app_id / feishu_app_secret …
            try:
                a = str(s["feishu_app_id"])
                b = str(s["feishu_app_secret"])
                c = str(s["feishu_app_token"])
                d = str(s["feishu_table_id"])
                if a and b and c and d:
                    return {"app_id": a, "app_secret": b, "app_token": c, "table_id": d}
            except (KeyError, TypeError):
                pass
    except Exception:
        pass
    # 2) 环境变量（Glitch / Railway / 自建等）
    try:
        a = os.environ.get("FEISHU_APP_ID") or os.environ.get("feishu_app_id")
        b = os.environ.get("FEISHU_APP_SECRET") or os.environ.get("feishu_app_secret")
        c = os.environ.get("FEISHU_APP_TOKEN") or os.environ.get("feishu_app_token")
        d = os.environ.get("FEISHU_TABLE_ID") or os.environ.get("feishu_table_id")
        if a and b and c and d:
            return {"app_id": a, "app_secret": b, "app_token": c, "table_id": d}
    except Exception:
        pass
    return None


def _get_shein_from_secrets():
    """若部署时在后台配置了 SHEIN 图片转链密钥，则返回且不在页面展示。"""
    try:
        s = getattr(st, "secrets", None)
        if s is not None:
            try:
                shein = s["shein"]
                o = str(shein["open_key_id"])
                k = str(shein["secret_key"])
                if o and k:
                    return {"open_key_id": o, "secret_key": k}
            except (KeyError, TypeError):
                pass
            try:
                o = str(s["shein_open_key_id"])
                k = str(s["shein_secret_key"])
                if o and k:
                    return {"open_key_id": o, "secret_key": k}
            except (KeyError, TypeError):
                pass
    except Exception:
        pass
    try:
        o = os.environ.get("SHEIN_OPEN_KEY_ID") or os.environ.get("shein_open_key_id")
        k = os.environ.get("SHEIN_SECRET_KEY") or os.environ.get("shein_secret_key")
        if o and k:
            return {"open_key_id": o, "secret_key": k}
    except Exception:
        pass
    return None


# 优先使用部署后台配置的 Secrets（仅管理员可见，同事页面不展示）
_feishu_from_secrets = _get_feishu_from_secrets()
_shein_from_secrets = _get_shein_from_secrets()

# 若后台已配置 Secrets，自动在后台静默连接飞书，无需用户手动点击
if _feishu_from_secrets and not st.session_state.get("feishu_fields"):
    try:
        _auto_token = get_feishu_tenant_token(
            _feishu_from_secrets["app_id"], _feishu_from_secrets["app_secret"]
        )
        _auto_fields = fetch_feishu_bitable_fields(
            _auto_token, _feishu_from_secrets["app_token"], _feishu_from_secrets["table_id"]
        )
        st.session_state.feishu_fields = _auto_fields or []
        st.session_state.config_mapping["_feishu_auth"] = _feishu_from_secrets
    except Exception:
        pass

with st.sidebar:
    # Secrets 已配置时完全隐藏飞书配置区，同事看不到任何凭证入口
    if not _feishu_from_secrets:
        st.header("🔗 飞书多维表格配置")
        st.caption("填入凭证以动态获取最新字段（输入一次自动记住）")
        _fs_cache = _load_feishu_cache()
        saved_fs = _fs_cache or st.session_state.config_mapping.get("_feishu_auth", {})
        fs_app_id = st.text_input("App ID (自建应用)", value=saved_fs.get("app_id", ""))
        fs_app_secret = st.text_input("App Secret", type="password", value=saved_fs.get("app_secret", ""))
        fs_app_token = st.text_input("App Token (多维表格)", value=saved_fs.get("app_token", ""))
        fs_table_id = st.text_input("Table ID (数据表)", value=saved_fs.get("table_id", ""))
    else:
        # 使用后台 Secrets，前端不显示任何飞书配置
        saved_fs = _feishu_from_secrets
        fs_app_id = saved_fs["app_id"]
        fs_app_secret = saved_fs["app_secret"]
        fs_app_token = saved_fs["app_token"]
        fs_table_id = saved_fs["table_id"]

    # Secrets 已配置时隐藏手动连接按钮和字段查看器（后台已自动连接）
    if not _feishu_from_secrets:
        if st.button("🔌 测试连接并获取飞书字段", use_container_width=True):
            if not all([fs_app_id, fs_app_secret, fs_app_token, fs_table_id]):
                st.error("请完整填写上方 4 个飞书凭证字段")
            else:
                with st.spinner("正在安全连接飞书开放平台..."):
                    try:
                        token = get_feishu_tenant_token(fs_app_id, fs_app_secret)
                        fields = fetch_feishu_bitable_fields(token, fs_app_token, fs_table_id)
                        if not fields:
                            st.warning("请求成功，但该表格中似乎没有任何字段/列！")
                            st.session_state.feishu_fields = []
                        else:
                            st.session_state.feishu_fields = fields
                            _auth_data = {
                                "app_id": fs_app_id,
                                "app_secret": fs_app_secret,
                                "app_token": fs_app_token,
                                "table_id": fs_table_id
                            }
                            st.session_state.config_mapping["_feishu_auth"] = _auth_data
                            _save_feishu_cache(_auth_data)
                            st.success(f"✅ 成功连接！已拉取到 {len(fields)} 个字段。凭据已自动保存。")
                    except requests.exceptions.HTTPError as he:
                        st.error(f"网络异常 (HTTP {he.response.status_code}): 请检查 ID 是否存在或格式正确。")
                    except Exception as e:
                        st.error(f"错误: {str(e)}")

        if st.session_state.feishu_fields:
            with st.expander(f"查看已缓存的 {len(st.session_state.feishu_fields)} 个字段"):
                st.write(st.session_state.feishu_fields)

    # ── SHEIN 图片转链配置（暂时隐藏，需要时取消注释恢复） ──
    # st.markdown("---")
    # st.header("🖼️ SHEIN 图片转链")
    # if _shein_from_secrets:
    #     st.caption("SHEIN 开放平台密钥已由管理员在后台配置，可直接使用（侧栏不展示）。")
    #     _shein_open_key_id = _shein_from_secrets["open_key_id"]
    #     _shein_secret_key = _shein_from_secrets["secret_key"]
    # else:
    #     st.caption("填写后自动保存到本地；后续可改为后台 Secrets / 环境变量以隐藏此处。第二步「本地图片转直链」列优先直传 SHEIN，失败时图床+转链。")
    #     _shein_saved = _load_shein_cache() or st.session_state.get("shein_auth", {})
    #     if _shein_saved and not st.session_state.get("shein_auth"):
    #         st.session_state.shein_auth = _shein_saved
    #     _shein_open_key_id = st.text_input("Open Key ID", value=_shein_saved.get("open_key_id", ""), key="shein_open_key_id")
    #     _shein_secret_key = st.text_input("Secret Key", type="password", value=_shein_saved.get("secret_key", ""), key="shein_secret_key")
    #     if _shein_open_key_id or _shein_secret_key:
    #         st.session_state.shein_auth = {"open_key_id": _shein_open_key_id, "secret_key": _shein_secret_key}
    #         if _shein_open_key_id and _shein_secret_key:
    #             _save_shein_cache(st.session_state.shein_auth)

# ---------- 主面板：页面切换（侧边栏底部） ----------
with st.sidebar:
    st.markdown("---")
    if 'current_step' not in st.session_state:
        st.session_state.current_step = "step1"
    _step_labels = {"step1": "📝 第一步：模板配置与保存", "step2": "🚀 第二步：生成上传文件"}
    _step_sel = st.radio("选择操作步骤", list(_step_labels.keys()),
                         format_func=lambda x: _step_labels[x], label_visibility="collapsed")
    st.session_state.current_step = _step_sel

# ================================================================
# ██ 第一步：模板配置与保存 ██
# ================================================================
if st.session_state.current_step == "step1":
    st.title("📝 第一步：模板配置与保存")
    st.markdown("上传官方模板，配置固定属性与飞书绑定关系，保存为模板文件。")
    st.subheader("📦 1. 快速加载已有模板 (可选)")
    saved_templates = get_saved_templates()
    
    col_sel, col_btn = st.columns([3, 1])
    with col_sel:
        selected_template = st.selectbox(
            "选择已保存的配置模板反填", 
            options=["-- 未选择 --"] + saved_templates,
            help="选择一个历史保存的 JSON 模板，将直接反填下方所有映射规则。"
        )
    with col_btn:
        st.write("") # 占位对齐
        st.write("")
        if st.button("🔄 一键加载模板"):
            if selected_template != "-- 未选择 --":
                try:
                    loaded_data = load_template(selected_template)
                    if not loaded_data:
                        st.warning("⚠️ 模板文件为空或不存在。")
                    else:
                        st.session_state.config_mapping = loaded_data
                        st.session_state.loaded_template_name = selected_template
                        # ── 清除所有 widget key 缓存（解决 Streamlit Widget Key 不随 index 更新的问题）──
                        _widget_prefixes = (
                            "type_", "val_", "disabled_val_",
                            "zone_attr_", "zone_qty_", "zone_currency_", "zone_date_",
                            "feishu_bind_",
                            "attr_req_", "attr_opt_",
                        )
                        _keys_to_del = [
                            k for k in list(st.session_state.keys())
                            if any(k.startswith(p) for p in _widget_prefixes)
                        ]
                        for _k in _keys_to_del:
                            del st.session_state[_k]
                        st.success(f"✅ 模板 [{selected_template}] 加载成功！规格属性、飞书绑定、件数、货币、日期等已选配置将直接反填到下方。")
                        if "_feishu_auth" in loaded_data:
                            st.info("检测到该模板内含飞书凭证缓存，正在左侧侧边栏自动填充。请手动点击『测试连接并获取飞书字段』以激活字段列表！")
                        st.rerun()  # 强制刷新页面，使已加载配置立即显示在下方选项区域
                except Exception as e:
                    st.error(f"❌ 加载模板失败: {str(e)}")
            else:
                st.warning("请先从左侧选择一个有效模板。")

    # 删除模板：每个模板一行，右侧可点击删除
    if saved_templates:
        st.caption("删除不需要的配置模板（点击对应行的删除即可）")
        for _tpl_name in saved_templates:
            _row1, _row2 = st.columns([4, 1])
            with _row1:
                st.text(_tpl_name)
            with _row2:
                if st.button("🗑️ 删除", key=f"del_tpl_{_tpl_name}", type="secondary"):
                    ok, err = delete_template(_tpl_name)
                    if ok:
                        if st.session_state.get("loaded_template_name") == _tpl_name:
                            st.session_state.loaded_template_name = ""
                        st.success(f"✅ 已删除：{_tpl_name}")
                        st.rerun()
                    else:
                        st.error(f"❌ 删除失败：{err}")

    st.markdown("---")

    # ---------- 上传区与解析逻辑 ----------
    st.subheader("📤 2. 上传 SHEIN 官方商品模板")
    uploaded_file = st.file_uploader("拖拽或选择 .xlsx 文件，程序将自动读取表头进行映射配置", type=['xlsx'])

    if uploaded_file:
        # 保存原始 Excel 字节供第二步复用
        uploaded_file.seek(0)
        st.session_state['uploaded_excel_bytes'] = uploaded_file.read()
        uploaded_file.seek(0)
        try:
            # 读取指定 Sheet 和 表头
            # 注意: 业务要求为第2个 sheet (索引1), 第2行是列名，第6行是规则
            xlsx = pd.ExcelFile(uploaded_file, engine='openpyxl')
            target_sheet = xlsx.sheet_names[1] if len(xlsx.sheet_names) > 1 else xlsx.sheet_names[0]
        
            # 提取表头和规则 (前10行足矣)
            df_meta = pd.read_excel(uploaded_file, sheet_name=target_sheet, header=None, nrows=10, engine='openpyxl')
        
            raw_columns_row = df_meta.iloc[1]
            raw_required_row = df_meta.iloc[3]  # 第4行："必填" / "非必填"
            raw_rules_row = df_meta.iloc[5]
        
            clean_columns = []
            excel_rules_dict = {}
            excel_required_dict = {}  # {col_name: "必填" / "选填" / ...}
        
            for col_name, req_val, rule_val in zip(raw_columns_row, raw_required_row, raw_rules_row):
                col_str = str(col_name).strip()
                if col_str and not col_str.startswith('Unnamed') and col_str != 'nan' \
                   and not col_str.startswith('请先选择') and not col_str.startswith('请选择'):
                    clean_columns.append(col_str)
                    excel_rules_dict[col_str] = str(rule_val).strip()
                    excel_required_dict[col_str] = str(req_val).strip() if pd.notna(req_val) else ""
        
            if not clean_columns:
                st.error("❌ 未能在目标 Sheet 的第 2 行找到有效的表头列名，请检查 Excel 格式！")
                st.stop()
            
            st.session_state.excel_columns = clean_columns
            st.session_state.excel_rules_dict = excel_rules_dict
            st.session_state.excel_required_dict = excel_required_dict
        
            # 提取选项字典 (有效属性值 & 有效品牌列表)
            options_dict = {}
        
            # 1. 尝试解析【有效品牌列表】
            brand_sheet_name = next((s for s in xlsx.sheet_names if "有效品牌" in s), None)
            if brand_sheet_name:
                try:
                    df_brands = pd.read_excel(uploaded_file, sheet_name=brand_sheet_name, engine='openpyxl')
                    # 寻找包含品牌的列名
                    brand_col = next((col for col in df_brands.columns if col and ("品牌" in str(col) or "Brand" in str(col))), None)
                    if brand_col:
                        brand_list = df_brands[brand_col].dropna().astype(str).str.strip().tolist()
                        brand_list = [b for b in brand_list if b and b.lower() != 'nan']
                        if brand_list:
                            options_dict["品牌"] = brand_list
                except Exception as e:
                    st.warning(f"解析【有效品牌列表】出现警告 (可忽略): {e}")

            # 2. 尝试解析【有效属性值】—— 横向(Row-based)数据结构
            #    前2行废话，第3行表头，每行一个属性：Key在"属性"列，选项从第6列横向展开
            attr_sheet_name = next((s for s in xlsx.sheet_names if "有效属性" in s), None)
            _attr_debug = {}  # 诊断信息
            if attr_sheet_name:
                try:
                    # 先用 header=None 读取前几行，看看真实结构
                    df_raw_peek = pd.read_excel(
                        uploaded_file, sheet_name=attr_sheet_name,
                        header=None, nrows=8, engine='openpyxl'
                    )
                    _attr_debug['raw_peek_shape'] = str(df_raw_peek.shape)
                    _attr_debug['row0'] = [str(x) for x in df_raw_peek.iloc[0].tolist()[:10]]
                    _attr_debug['row1'] = [str(x) for x in df_raw_peek.iloc[1].tolist()[:10]]
                    _attr_debug['row2'] = [str(x) for x in df_raw_peek.iloc[2].tolist()[:10]]
                    if len(df_raw_peek) > 3:
                        _attr_debug['row3'] = [str(x) for x in df_raw_peek.iloc[3].tolist()[:10]]
                    if len(df_raw_peek) > 4:
                        _attr_debug['row4'] = [str(x) for x in df_raw_peek.iloc[4].tolist()[:10]]

                    df_attr = pd.read_excel(
                        uploaded_file, sheet_name=attr_sheet_name,
                        header=1, engine='openpyxl'  # row0=废话, row1=真正表头
                    )

                    _attr_debug['columns'] = [str(c) for c in df_attr.columns.tolist()[:15]]
                    _attr_debug['shape'] = str(df_attr.shape)

                    # 定位"属性"列：优先精确匹配，避免"属性类型"误命中
                    key_col_idx = None
                    for ci, c in enumerate(df_attr.columns):
                        if str(c).strip() == "属性" or str(c).strip() == "属性名" or str(c).strip() == "属性名称":
                            key_col_idx = ci
                            break
                    if key_col_idx is None:
                        # fallback: 包含"属性"但排除"属性类型"/"属性值"
                        for ci, c in enumerate(df_attr.columns):
                            cs = str(c).strip()
                            if "属性" in cs and "类型" not in cs and "值" not in cs and "必填" not in cs and "重要" not in cs:
                                key_col_idx = ci
                                break
                    if key_col_idx is None:
                        key_col_idx = 2  # 最终默认第3列
                    _attr_debug['key_col_idx'] = key_col_idx
                    _attr_debug['key_col_name'] = str(df_attr.columns[key_col_idx])

                    # 逐行横向提取选项
                    for _, row in df_attr.iterrows():
                        key_val = row.iloc[key_col_idx]
                        if pd.isna(key_val) or not str(key_val).strip():
                            continue
                        key_str = str(key_val).strip()

                        raw_opts = row.iloc[5:]
                        clean_opts = []
                        for v in raw_opts:
                            if pd.notna(v):
                                vs = str(v).strip()
                                if vs and vs.lower() != 'nan':
                                    clean_opts.append(vs)
                        if clean_opts:
                            options_dict[key_str] = clean_opts

                except Exception as e:
                    st.warning(f"解析【有效属性值】出现警告 (可忽略): {e}")
                    _attr_debug['error'] = str(e)

            # 【诊断面板】有效属性值解析结果
            with st.expander("🔬 有效属性值字典诊断（调试用）", expanded=False):
                if _attr_debug:
                    for dk, dv in _attr_debug.items():
                        st.write(f"**{dk}**: {dv}")
                st.markdown("---")
                attr_keys = [k for k in options_dict.keys() if k != "品牌"]
                st.write(f"**attr_dict 共 {len(attr_keys)} 个 Key**：{attr_keys[:20]}")
                if attr_keys:
                    first_k = attr_keys[0]
                    st.write(f"**示例 '{first_k}'**: {options_dict[first_k][:5]}...")


            st.session_state.dv_options_dict = {}

            # 3. 深度扫描主表原生下拉菜单 (Data Validation) —— 完整重构版
            def _extract_range_values(ws_ref, range_str):
                """从指定 Sheet 的指定范围提取所有非空字符串值。"""
                collected = []
                try:
                    # ✅ Bug Fix 1: 去掉 $ 符号，openpyxl 不支持 $F$6:$G$6 这种写法
                    clean_range = range_str.replace('$', '')
                    cells = ws_ref[clean_range]
                    if isinstance(cells, tuple):
                        for row_or_col in cells:
                            if isinstance(row_or_col, tuple):
                                for cell in row_or_col:
                                    if cell.value is not None:
                                        v = str(cell.value).strip()
                                        if v and v.lower() != 'nan':
                                            collected.append(v)
                            else:
                                if row_or_col.value is not None:
                                    v = str(row_or_col.value).strip()
                                    if v and v.lower() != 'nan':
                                        collected.append(v)
                    else:
                        if cells.value is not None:
                            v = str(cells.value).strip()
                            if v and v.lower() != 'nan':
                                collected.append(v)
                except Exception:
                    pass
                return collected

            def _parse_sheet_ref(ref_str, wb, target_sheet):
                """解析 'SheetName'!Range 或 SheetName!Range 格式，返回 (sheet_name, range_str)。"""
                if '!' not in ref_str:
                    return target_sheet, ref_str
                bang_idx = ref_str.index('!')
                sheet_name = ref_str[:bang_idx]
                rng = ref_str[bang_idx + 1:]
                # 去掉 Sheet 名两侧的引号
                if len(sheet_name) >= 2 and sheet_name[0] in ("'", '"') and sheet_name[-1] == sheet_name[0]:
                    sheet_name = sheet_name[1:-1]
                return sheet_name, rng

            def _resolve_sqref_to_colnames(dv, raw_columns_row):
                """根据 dv.sqref 解析出所有关联的列名列表。"""
                from openpyxl.utils.cell import coordinate_to_tuple
                names = []
                if not dv.sqref:
                    return names
                for rng in str(dv.sqref).split():
                    start_cell = rng.split(':')[0]
                    col_letter = ''.join(c for c in start_cell if c.isalpha())
                    if not col_letter:
                        continue
                    try:
                        _, start_col = coordinate_to_tuple(col_letter + '1')
                    except Exception:
                        continue
                    col_idx = start_col - 1
                    if 0 <= col_idx < len(raw_columns_row):
                        n = str(raw_columns_row.iloc[col_idx]).strip()
                        if n and not n.startswith('Unnamed') and n != 'nan':
                            names.append(n)
                return names

            def _map_dv_cols(dv, raw_columns_row, clean_vals, dv_dict, opt_dict):
                """根据 dv.sqref 将 clean_vals 写入对应列名字典（情况A和B共用）。"""
                if not clean_vals or not dv.sqref:
                    return
                for col_name_mapped in _resolve_sqref_to_colnames(dv, raw_columns_row):
                    dv_dict[col_name_mapped] = clean_vals
                    opt_dict[col_name_mapped] = clean_vals

            def _sanitize_for_named_range(val):
                """复现 Excel SUBSTITUTE 链：把特殊字符全部替换为下划线。"""
                special = ' -.()\'[]{}*+?？^$|\\/:;,<>!@#%&="~`'
                result = str(val)
                for ch in special:
                    result = result.replace(ch, '_')
                return result

            def _parse_indirect_cascade(f1_str, dv, raw_columns_row, wb, dv_local, options_dict):
                """解析 INDIRECT 级联公式，返回 cascade_config 或 None。"""
                import re
                from openpyxl.utils.cell import coordinate_to_tuple

                # 1. 提取父列字母（如 $CC7 → CC）
                parent_ref = re.search(r'\$([A-Z]+)\$?\d+', f1_str)
                if not parent_ref:
                    return None
                parent_letter = parent_ref.group(1)
                try:
                    _, p_col_num = coordinate_to_tuple(parent_letter + '1')
                except Exception:
                    return None
                p_idx = p_col_num - 1
                if p_idx < 0 or p_idx >= len(raw_columns_row):
                    return None
                parent_col = str(raw_columns_row.iloc[p_idx]).strip()
                if not parent_col or parent_col.startswith('Unnamed') or parent_col == 'nan':
                    return None

                # 2. 提取命名范围前缀（如 "CAT_15604_"）
                prefix_match = re.search(r'"([A-Za-z0-9_]+_)"\s*&', f1_str)
                if not prefix_match:
                    return None
                prefix = prefix_match.group(1)

                # 3. 获取父列已解析的选项
                parent_options = dv_local.get(parent_col, options_dict.get(parent_col, []))
                if not parent_options:
                    return None

                # 4. 获取子列列名
                child_cols = _resolve_sqref_to_colnames(dv, raw_columns_row)
                if not child_cols:
                    return None

                # 5. 对每个父选项，计算命名范围名并查找
                cascade_mapping = {}  # {parent_val: [child_vals]}
                for pv in parent_options:
                    sanitized = _sanitize_for_named_range(pv)
                    nr_name = prefix + sanitized
                    # wb.defined_names 支持按名查找
                    defn = wb.defined_names.get(nr_name)
                    if defn is None:
                        continue
                    try:
                        for dest_sheet, cell_range in defn.destinations:
                            if dest_sheet in wb.sheetnames:
                                vals = _extract_range_values(wb[dest_sheet], cell_range)
                                if vals:
                                    cascade_mapping[pv] = vals
                                    break
                    except Exception:
                        pass

                if not cascade_mapping:
                    return None

                return {
                    'parent_col': parent_col,
                    'child_cols': child_cols,
                    'mapping': cascade_mapping
                }

            try:
                import openpyxl
                uploaded_file.seek(0)
                wb = openpyxl.load_workbook(uploaded_file, data_only=False)

                if target_sheet in wb.sheetnames:
                    ws = wb[target_sheet]
                    dv_local = {}
                    dv_fallback_cols = []
                    cascade_configs = []  # 级联关系列表

                    dvs = ws.data_validations.dataValidation if hasattr(ws, 'data_validations') else []
                    for dv in dvs:
                        if dv.type != 'list' or not dv.formula1:
                            continue

                        f1_str = str(dv.formula1).strip()
                        clean_vals = []

                        # ── 情况 A：带引号的字面量字符串 "选项1,选项2" ─────────────────
                        if (f1_str.startswith('"') and f1_str.endswith('"')) or \
                           (f1_str.startswith("'") and f1_str.endswith("'") and '!' not in f1_str):
                            inner = f1_str[1:-1]
                            clean_vals = [v.strip() for v in inner.replace('，', ',').split(',') if v.strip()]

                        # ── 情况 A2：无引号但含逗号的字面量 选项1,选项2 ────────────────
                        elif not f1_str.startswith('=') and '!' not in f1_str and (',' in f1_str or '，' in f1_str):
                            clean_vals = [v.strip() for v in f1_str.replace('，', ',').split(',') if v.strip()]

                        # ── INDIRECT 动态公式 → 优先级联解析，fallback 模糊匹配 ───
                        elif 'INDIRECT(' in f1_str.upper():
                            cascade_info = _parse_indirect_cascade(
                                f1_str, dv, raw_columns_row, wb, dv_local, options_dict
                            )
                            if cascade_info:
                                cascade_configs.append(cascade_info)
                                # 把所有子选项合并写入 dv_local，确保推断引擎识别为固定单选
                                all_child_vals = []
                                for vals in cascade_info['mapping'].values():
                                    for v in vals:
                                        if v not in all_child_vals:
                                            all_child_vals.append(v)
                                for cc in cascade_info['child_cols']:
                                    dv_local[cc] = all_child_vals
                                    options_dict[cc] = all_child_vals
                            else:
                                # 级联解析失败 → 降级到辅助 Sheet 模糊匹配
                                affected_cols = _resolve_sqref_to_colnames(dv, raw_columns_row)
                                for ac in affected_cols:
                                    matched_vals = []
                                    for dict_key, dict_vals in options_dict.items():
                                        if dict_key in ac or ac in dict_key:
                                            matched_vals = dict_vals
                                            break
                                    if matched_vals:
                                        dv_local[ac] = matched_vals
                                        options_dict[ac] = matched_vals
                                    else:
                                        dv_fallback_cols.append(ac)

                        # ── 情况 B：= 开头的公式引用 =Sheet!Range ────────────────────
                        elif f1_str.startswith('='):
                            ref_str = f1_str[1:]
                            sheet_name, ref_range = _parse_sheet_ref(ref_str, wb, target_sheet)
                            if sheet_name in wb.sheetnames:
                                clean_vals = _extract_range_values(wb[sheet_name], ref_range)
                                # 品牌 Sheet 兜底：若范围读不到值，直接取第一列
                                if '品牌' in sheet_name and not clean_vals:
                                    for row in wb[sheet_name].iter_rows(min_col=1, max_col=1, values_only=True):
                                        v = str(row[0]).strip() if row[0] is not None else ''
                                        if v and v.lower() != 'nan':
                                            clean_vals.append(v)

                        # ── ✅ Bug Fix 2：无 = 前缀的裸 Sheet 引用 Sheet!Range ─────────
                        elif '!' in f1_str:
                            sheet_name, ref_range = _parse_sheet_ref(f1_str, wb, target_sheet)
                            if sheet_name in wb.sheetnames:
                                clean_vals = _extract_range_values(wb[sheet_name], ref_range)

                        # ── 共用：clean_vals → 列名字典 ──────────────────────────────
                        _map_dv_cols(dv, raw_columns_row, clean_vals, dv_local, options_dict)

                    st.session_state.dv_options_dict = dv_local
                    st.session_state.dv_fallback_cols = dv_fallback_cols
                    st.session_state.cascade_configs = cascade_configs

            except Exception as e:
                st.warning(f"解析主表原生下拉菜单 (Data Validation) 出现警告：{e}")



            st.session_state.options_dict = options_dict

            # 【诊断面板】供调试用，可随时折叠
            with st.expander("🔬 DV 规则诊断（调试用）", expanded=False):
                dv_debug = st.session_state.get('dv_options_dict', {})
                st.write(f"**成功解析的 DV 列数**: {len(dv_debug)}")
                if dv_debug:
                    for k, v in dv_debug.items():
                        st.write(f"- **{k}**: {len(v)} 个选项 → {v[:5]}{'...' if len(v)>5 else ''}")
                st.markdown("---")
                st.write("**所有原始 DV 规则（formula1 原文）**：")
                try:
                    uploaded_file.seek(0)
                    import openpyxl as _ox
                    _wb = _ox.load_workbook(uploaded_file, data_only=False)
                    if target_sheet in _wb.sheetnames:
                        _ws = _wb[target_sheet]
                        dvs_raw = _ws.data_validations.dataValidation if hasattr(_ws, 'data_validations') else []
                        for i, _dv in enumerate(dvs_raw):
                            st.code(f"[{i}] type={_dv.type}  sqref={_dv.sqref}\n     formula1={_dv.formula1}")
                except Exception as diag_e:
                    st.warning(f"诊断读取失败: {diag_e}")

            st.success(f"✅ 成功从 `{target_sheet}` 提取到 {len(clean_columns)} 个有效列名，并抓取了填写规则！")

        
        except Exception as e:
            st.error(f"❌ 解析 Excel 文件失败，错误详情: {str(e)}")
            st.stop()

    # ---------- 动态映射配置区（左右双列布局）---------- 
    if st.session_state.excel_columns:
        if 'config_mapping' not in st.session_state:
            st.session_state.config_mapping = {}
        
        options_dict = st.session_state.get('options_dict', {})
        excel_rules_dict = st.session_state.get('excel_rules_dict', {})

        # ── 分组：集合 A（核心）vs 集合 B（辅助） ──────────────────────────
        AUX_KEYWORDS = ["视频", "法语", "德语", "西班牙语", "葡萄牙语", "意大利语",
                        "泰语", "阿拉伯语", "日语", "韩语", "越南语", "马来语",
                        "商品名称-", "商品卖点-", "商品描述-", "多语言"]
    
        group_a_cols = []  # 核心
        group_b_cols = []  # 辅助
        for col_name in st.session_state.excel_columns:
            col_upper = str(col_name).upper()
            if any(kw in col_name for kw in AUX_KEYWORDS):
                group_b_cols.append(col_name)
            else:
                group_a_cols.append(col_name)

        # ── 共用渲染函数 ──────────────────────────────────────────────
        def _render_col_config(col_name, idx, new_mapping_ref):
            """渲染单列的类型选择 + 值输入组件，返回 col_config dict。"""
            rule_str = excel_rules_dict.get(col_name, "")
            prev_config = st.session_state.config_mapping.get(col_name, None)
        
            # --- 智能推断引擎（六重优先级） ---
            guessed_type = OP_IGNORE
            rule_upper = str(rule_str).upper()
            dv_options_dict = st.session_state.get('dv_options_dict', {})
            dv_fallback_cols = st.session_state.get('dv_fallback_cols', [])
        
            # 获取当前列在全局列表中的索引
            _all_cols = st.session_state.excel_columns
            _col_idx = _all_cols.index(col_name) if col_name in _all_cols else -1
        
            # 找"细节图"和"规格1"的位置，用于区间推断
            _zone_start = -1  # 细节图
            _zone_end = len(_all_cols)  # 规格1
            for _zi, _zc in enumerate(_all_cols):
                if "细节图" in _zc and _zone_start < 0:
                    _zone_start = _zi
                if ("规格1" in _zc or "主规格" in _zc) and _zi > _zone_start >= 0:
                    _zone_end = _zi
                    break
        
            # 【P0：核心动态列 → 运行态手动填写（多语言商品名称除外）】
            if "多语言" not in col_name and any(kw in col_name for kw in MANUAL_KEYWORDS):
                guessed_type = OP_MANUAL
            # 【P1：品牌列 → 默认固定单选，运营可手动切飞书】
            elif "品牌" in col_name:
                guessed_type = OP_FIXED_SINGLE
            # 【P2：DV 选项存在 → 固定单选】
            elif col_name in dv_options_dict:
                guessed_type = OP_FIXED_SINGLE
            # 【P2.5：INDIRECT 降级失败 → 飞书兜底】
            elif col_name in dv_fallback_cols:
                guessed_type = OP_FEISHU
            # 【P3：第六行文字规则关键词】
            elif any(kw in rule_upper for kw in ["图片", "JPG", "像素", "尺寸"]) or col_name.rstrip(')').rstrip('0123456789').endswith("图"):
                guessed_type = OP_IMAGE
            elif "多选" in rule_upper:
                guessed_type = OP_FIXED_MULTI
            elif "单选" in rule_upper:
                guessed_type = OP_FIXED_SINGLE
            # 【P4：辅助信息列 → 忽略】
            elif any(kw in col_name for kw in AUX_KEYWORDS):
                guessed_type = OP_IGNORE
            # 【P5 兜底：细节图~规格1 区间 → 飞书；其它 → 飞书】
            else:
                guessed_type = OP_FEISHU
        
            # 强行映射覆盖
            type_str_map = {
                "ignore": OP_IGNORE, "fixed_single": OP_FIXED_SINGLE,
                "fixed_multi": OP_FIXED_MULTI, "fixed_text": OP_FIXED_TEXT,
                "feishu": OP_FEISHU, "image": OP_IMAGE, "manual": OP_MANUAL
            }
            reverse_type_map = {
                OP_IGNORE: "ignore", OP_FIXED_SINGLE: "fixed_single",
                OP_FIXED_MULTI: "fixed_multi", OP_FIXED_TEXT: "fixed_text",
                OP_FEISHU: "feishu", OP_IMAGE: "image", OP_MANUAL: "manual"
            }
            if prev_config:
                if prev_config.get("type") == "fixed":
                    default_op_str = guessed_type
                else:
                    default_op_str = type_str_map.get(prev_config.get("type"), guessed_type)
            else:
                default_op_str = guessed_type
        
            default_index = PROCESSING_OPTIONS.index(default_op_str)
        
            # 拼接必填标记
            req_dict = st.session_state.get('excel_required_dict', {})
            req_text = req_dict.get(col_name, '')
            req_badge = ' 🔴必填' if ('必填' in req_text and '非必填' not in req_text) else ' ⚪选填'
        
            # 初始化该列在 mapping 中的结构 (如果尚未存在)
            if col_name not in st.session_state.config_mapping:
                st.session_state.config_mapping[col_name] = {"type": reverse_type_map[default_op_str]}

            # 统一同步回调：将当前 widget 的值立刻刷新到 persistent mapping 中
            def _sync_val(key_name, field_name):
                st.session_state.config_mapping[col_name][field_name] = st.session_state[key_name]
                if field_name == "value" and st.session_state[key_name] in ("-- 请选择 --", "-- 不填 --"):
                    st.session_state.config_mapping[col_name][field_name] = ""
                # 如果是多选组件，将列表转为 * 拼接的字符串
                if isinstance(st.session_state[key_name], list):
                    st.session_state.config_mapping[col_name][field_name] = "*".join(st.session_state[key_name])
            
            def _sync_type(key_name):
                t_val = st.session_state[key_name]
                st.session_state.config_mapping[col_name]["type"] = reverse_type_map[t_val]

            with st.expander(f"📌 {idx+1}. {_display_name(col_name)}{req_badge}  |  🤖[{guessed_type}]", expanded=False):
                st.caption(f"**规则**: {rule_str if rule_str and str(rule_str) != 'nan' else '—'}")
            
                col_type, col_val = st.columns([1, 2])
            
                type_wk = f"type_{idx}_{col_name}"
                with col_type:
                    selected_type = st.selectbox(
                        "处理类型", options=PROCESSING_OPTIONS,
                        index=default_index, key=type_wk,
                        on_change=_sync_type, args=(type_wk,)
                    )
            
                # 同步回 new_mapping_ref (保持原来引用有效)
                new_mapping_ref[col_name] = st.session_state.config_mapping[col_name]
            
                with col_val:
                    if selected_type == OP_FIXED_SINGLE:
                        dv_d = st.session_state.get('dv_options_dict', {})
                        prev_val = st.session_state.config_mapping[col_name].get("value", "")
                        is_required = '必填' in req_text and '非必填' not in req_text
                        EMPTY_PLACEHOLDER = "-- 不填 --"
                        SELECT_PLACEHOLDER = "-- 请选择 --"
                        def _inject_empty(opts):
                            """在选项前插入空值占位符"""
                            if is_required:
                                return [SELECT_PLACEHOLDER] + list(opts)
                            return [EMPTY_PLACEHOLDER] + list(opts)
                        
                        # 级联检测
                        _cascade_hit = None
                        for _cc in st.session_state.get('cascade_configs', []):
                            if col_name in _cc['child_cols']:
                                _cascade_hit = _cc
                                break

                        if _cascade_hit:
                            parent_col = _cascade_hit['parent_col']
                            cascade_map = _cascade_hit['mapping']
                            parent_current = st.session_state.config_mapping.get(parent_col, {}).get('value', '')
                            if parent_current and parent_current in cascade_map:
                                child_opts = _inject_empty(cascade_map[parent_current])
                                sel_idx = child_opts.index(prev_val) if prev_val in child_opts else 0
                                _wk = f"val_{idx}_{col_name}"
                                st.selectbox(
                                    f"🔗 级联（{parent_col}={parent_current}）",
                                    options=child_opts, index=sel_idx,
                                    key=_wk, on_change=_sync_val, args=(_wk, "value")
                                )
                            elif parent_current:
                                st.warning(f"⚠️ '{parent_current}' 无子选项")
                                _wk = f"val_{idx}_{col_name}"
                                st.text_input(
                                    f"手动输入", value=prev_val, key=_wk,
                                    on_change=_sync_val, args=(_wk, "value")
                                )
                            else:
                                st.info(f"⬆️ 请先选择【{parent_col}】")
                                st.text_input("等待父列选择", value=prev_val, disabled=True, key=f"val_{idx}_{col_name}")

                        elif col_name in dv_d:
                            val_options = dv_d[col_name]
                            if val_options:
                                display_opts = _inject_empty(val_options)
                                sel_idx = display_opts.index(prev_val) if prev_val in display_opts else 0
                                _wk = f"val_{idx}_{col_name}"
                                st.selectbox(
                                    f"🥇 内置下拉选项", options=display_opts, index=sel_idx,
                                    key=_wk, on_change=_sync_val, args=(_wk, "value")
                                )
                            else:
                                _wk = f"val_{idx}_{col_name}"
                                st.text_input(
                                    f"⚠️ 未抓取到选项", value=prev_val, key=_wk,
                                    on_change=_sync_val, args=(_wk, "value")
                                )
                        else:
                            val_options = []
                            matched_key = ""
                            for dict_key, dict_vals in options_dict.items():
                                if dict_key in col_name or col_name in dict_key:
                                    val_options = dict_vals
                                    matched_key = dict_key
                                    break
                            if val_options:
                                display_opts = _inject_empty(val_options)
                                sel_idx = display_opts.index(prev_val) if prev_val in display_opts else 0
                                _wk = f"val_{idx}_{col_name}"
                                st.selectbox(
                                    f"从【{matched_key}】属性池", options=display_opts, index=sel_idx,
                                    key=_wk, on_change=_sync_val, args=(_wk, "value")
                                )
                            else:
                                _wk = f"val_{idx}_{col_name}"
                                st.text_input(
                                    f"手动输入", value=prev_val, key=_wk,
                                    on_change=_sync_val, args=(_wk, "value")
                                )

                    elif selected_type == OP_FIXED_MULTI:
                        val_options = []
                        matched_key = ""
                        for dict_key, dict_vals in options_dict.items():
                            if dict_key in col_name or col_name in dict_key:
                                val_options = dict_vals
                                matched_key = dict_key
                                break
                        prev_val_str = st.session_state.config_mapping[col_name].get("value", "")
                        prev_val_list = [v.strip() for v in prev_val_str.split("*")] if prev_val_str else []
                        if val_options:
                            valid_prev = [v for v in prev_val_list if v in val_options]
                            _wk = f"val_{idx}_{col_name}"
                            st.multiselect(
                                f"从【{matched_key}】多选", options=val_options, default=valid_prev,
                                key=_wk, on_change=_sync_val, args=(_wk, "value")
                            )
                        else:
                            st.caption(f"⚠️ 未匹配到 [{col_name}] 的选项")
                            _wk = f"val_{idx}_{col_name}"
                            st.text_input(
                                f"手动输入 (多值用 * 隔开)", value=prev_val_str,
                                key=_wk, on_change=_sync_val, args=(_wk, "value")
                            )
                        
                    elif selected_type == OP_FIXED_TEXT:
                        prev_val = st.session_state.config_mapping[col_name].get("value", "")
                        _wk = f"val_{idx}_{col_name}"
                        st.text_input(f"固定文本", value=prev_val, key=_wk, on_change=_sync_val, args=(_wk, "value"))
                
                    elif selected_type == OP_FEISHU:
                        is_feishu_ready = len(st.session_state.feishu_fields) > 0
                        if is_feishu_ready:
                            feishu_options = st.session_state.feishu_fields
                            prev_feishu_key = st.session_state.config_mapping[col_name].get("feishu_key", "")
                            feishu_index = feishu_options.index(prev_feishu_key) if prev_feishu_key in feishu_options else 0
                            _wk = f"val_{idx}_{col_name}"
                            st.selectbox(
                                "飞书字段", options=feishu_options, index=feishu_index,
                                key=_wk, on_change=_sync_val, args=(_wk, "feishu_key")
                            )
                        else:
                            st.selectbox(
                                "未检测到飞书连接",
                                options=["🚫 请先在侧边栏连接飞书"],
                                disabled=True, key=f"disabled_val_{idx}_{col_name}"
                            )
                    
                    elif selected_type == OP_IMAGE:
                        st.info("🖼️ 运行态将自动接收图片直链")
                    
                    elif selected_type == OP_MANUAL:
                        st.caption("✍️ 该列将在运行态由操作员手动填写（货号/标题/描述等）")
                    
                    else:  # OP_IGNORE
                        st.caption("🙈 该列跳过")

        # 预填充 config_mapping：对所有列运行推断，但如果不渲染 UI 也要有默认值
        for col_name in st.session_state.excel_columns:
            if col_name.startswith('_'):
                continue
            if col_name not in st.session_state.config_mapping:
                # 运行推断引擎取默认值
                rule_str = excel_rules_dict.get(col_name, "")
                rule_upper = str(rule_str).upper()
                dv_opts = st.session_state.get('dv_options_dict', {})
                dv_fb = st.session_state.get('dv_fallback_cols', [])
                gt = OP_IGNORE
                if "多语言" not in col_name and any(kw in col_name for kw in MANUAL_KEYWORDS):
                    gt = OP_MANUAL
                elif "品牌" in col_name:
                    gt = OP_FIXED_SINGLE
                elif col_name in dv_opts:
                    gt = OP_FIXED_SINGLE
                elif col_name in dv_fb:
                    gt = OP_FEISHU
                elif any(kw in rule_upper for kw in ["图片", "JPG", "像素", "尺寸"]) or col_name.rstrip(')').rstrip('0123456789').endswith("图"):
                    gt = OP_IMAGE
                elif "多选" in rule_upper:
                    gt = OP_FIXED_MULTI
                elif "单选" in rule_upper:
                    gt = OP_FIXED_SINGLE
                elif any(kw in col_name for kw in AUX_KEYWORDS):
                    gt = OP_IGNORE
                else:
                    gt = OP_FEISHU
                r_map = {OP_IGNORE: "ignore", OP_FIXED_SINGLE: "fixed_single",
                         OP_FIXED_MULTI: "fixed_multi", OP_FIXED_TEXT: "fixed_text",
                         OP_FEISHU: "feishu", OP_IMAGE: "image", OP_MANUAL: "manual"}
                st.session_state.config_mapping[col_name] = {"type": r_map.get(gt, "ignore")}


        # ================================================================
        # ██ 规格与固定属性速览模块 ██
        # 映射规则（通用，适配不同模板）：
        # - 区间：Excel 中「规格3内容」之后 到 「视频」（含视频[shein-www] 等）之前的列
        # - 必填/选填：以 Excel 第 4 行（excel_required_dict）为准，必填展示在「必填属性」，选填折叠在「展开查看选填属性」
        # ================================================================
        st.markdown("---")
        st.header("📋 3. 规格与固定属性速览")
        st.caption("区间：规格3内容 → 视频（如 视频[shein-www]）之间的列；必填/选填以 Excel 第 4 行判断，必填直接展示，选填折叠。")

        mapping = st.session_state.get('config_mapping', {})
        req_dict = st.session_state.get('excel_required_dict', {})  # 来自 Excel 第 4 行
        opts_dict = st.session_state.get('options_dict', {})

        dv_d = st.session_state.get('dv_options_dict', {})
        EMPTY_PH = "-- 不填 --"

        # ── 区间：规格3内容 之后 ~ 视频 之前（含 视频[shein-www] 等，关键词「视频」）────
        all_cols = st.session_state.excel_columns
        range_start = 0
        range_end = len(all_cols)
        for i, c in enumerate(all_cols):
            if "规格3内容" in c:
                range_start = i + 1
        for i, c in enumerate(all_cols):
            if "视频" in c and i > range_start:
                range_end = i
                break

        attr_required = []
        attr_optional = []
        image_cols = []

        for i in range(range_start, range_end):
            col_name = all_cols[i]
            cfg = mapping.get(col_name, {})
            cfg_type = cfg.get("type", "ignore")
            if cfg_type == "ignore":
                continue
            if cfg_type == "image":
                image_cols.append(col_name)
                continue
            # 必填/选填：以 Excel 第 4 行对应单元格为准
            r = req_dict.get(col_name, '')
            is_req = '必填' in r and '非必填' not in r
            if is_req:
                attr_required.append(col_name)
            else:
                attr_optional.append(col_name)

        def _render_attr_widget(col_name, key_prefix):
            """根据配置类型渲染对应输入组件"""
            cfg = mapping.get(col_name, {})
            cfg_type = cfg.get("type", "ignore")
            label = _display_name(col_name)
            preset = cfg.get("value", "")
            r = req_dict.get(col_name, '')
            is_req = '必填' in r and '非必填' not in r

            SELECT_PH = "-- 请选择 --"

            if cfg_type == "fixed_single":
                opts = []
                if col_name in dv_d and dv_d[col_name]:
                    opts = list(dv_d[col_name])
                else:
                    for dk, dv_vals in opts_dict.items():
                        if dk in col_name or col_name in dk:
                            opts = list(dv_vals)
                            break
                if opts:
                    # 规格X内容：支持单选或自己填写（运营常用自定义文案），合并为一行：左下拉右输入
                    is_spec_content = "规格1内容" in col_name or "规格2内容" in col_name or "规格3内容" in col_name
                    MANUAL_FILL = "-- 自己填写 --"
                    display_opts = ([SELECT_PH, MANUAL_FILL] + opts) if is_req else ([EMPTY_PH, MANUAL_FILL] + opts)
                    if not is_spec_content:
                        display_opts = [SELECT_PH] + opts if is_req else [EMPTY_PH] + opts
                    widget_key = f"{key_prefix}_{col_name}"
                    custom_key = f"{widget_key}_custom"
                    # 若已保存的是自定义文案（不在选项里），默认选「自己填写」并带出文案
                    custom_default = preset if (preset and preset not in display_opts) else ""
                    if widget_key not in st.session_state and preset and preset in display_opts:
                        st.session_state[widget_key] = preset
                    elif widget_key not in st.session_state and custom_default:
                        st.session_state[widget_key] = MANUAL_FILL
                    sel_idx = display_opts.index(preset) if preset in display_opts else (display_opts.index(MANUAL_FILL) if custom_default else 0)
                    if is_spec_content:
                        st.caption(f"{label}  🔘单选/手填（选预设或右侧直接输入）")
                        _sc0, _sc1 = st.columns([1, 1.2])
                        with _sc0:
                            val = st.selectbox("预设", options=display_opts, index=sel_idx, key=widget_key, label_visibility="collapsed")
                        with _sc1:
                            custom_val = st.text_input("手填", value=custom_default, key=custom_key, placeholder="选「自己填写」或在此输入", label_visibility="collapsed")
                        if custom_val and str(custom_val).strip():
                            return str(custom_val).strip()
                        if val == SELECT_PH or val == EMPTY_PH or val == MANUAL_FILL:
                            return ""
                        return val
                    val = st.selectbox(f"{label}  🔘单选", options=display_opts, index=sel_idx, key=widget_key)
                    if val == SELECT_PH or val == EMPTY_PH:
                        return ""
                    return val
                else:
                    return st.text_input(f"{label}  🔘单选", value=preset, key=f"{key_prefix}_{col_name}")

            elif cfg_type == "fixed_multi":
                opts = []
                for dk, dv_vals in opts_dict.items():
                    if dk in col_name or col_name in dk:
                        opts = list(dv_vals)
                        break
                prev_list = [v.strip() for v in str(preset).split("*") if v.strip()] if preset else []
                if opts:
                    # 选填时多选与单选一致：默认不填，不沿用 preset，不写入 Excel
                    valid_prev = [v for v in prev_list if v in opts] if is_req else []
                    widget_key = f"{key_prefix}_{col_name}"
                    if widget_key not in st.session_state and valid_prev:
                        st.session_state[widget_key] = valid_prev
                    selected = st.multiselect(f"{label}  ☑️多选", options=opts, default=valid_prev, key=widget_key)
                    if not is_req:
                        st.caption("未选则不填，不写入 Excel")
                    return "*".join(selected)
                else:
                    return st.text_input(f"{label} (多值用*隔开)", value=preset, key=f"{key_prefix}_{col_name}")

            elif cfg_type == "fixed_text":
                return st.text_input(label, value=preset, key=f"{key_prefix}_{col_name}")

            elif cfg_type == "feishu":
                feishu_key = cfg.get("feishu_key", "")
                st.caption(f"🔗 飞书字段: `{feishu_key}`（运行态自动拉取）")
                return f"__FEISHU__{feishu_key}"

            elif cfg_type == "manual":
                st.caption(f"✍️ {label} — 运行态手动填写")
                return ""

            return ""

        # ── 必填属性直接展示（并持久化到 mapping 以便保存模板）──────
        if attr_required:
            st.caption(f"🔴 必填属性 ({len(attr_required)} 项)")
            for col_name in attr_required:
                val = _render_attr_widget(col_name, "attr_req")
                if val is not None and not str(val).startswith("__FEISHU__"):
                    mapping.setdefault(col_name, {})["value"] = str(val)

        # ── 选填属性折叠（并持久化到 mapping 以便保存模板）──────────
        if attr_optional:
            with st.expander(f"⚪ 展开查看选填属性 ({len(attr_optional)} 项)", expanded=False):
                for col_name in attr_optional:
                    val = _render_attr_widget(col_name, "attr_opt")
                    if val is not None and not str(val).startswith("__FEISHU__"):
                        mapping.setdefault(col_name, {})["value"] = str(val)


        # ================================================================
        # ██ 飞书字段绑定模块（细节图～规格1 之间的列，适配不同模板） ██
        # 映射规则：区间 = 最后一个细节图（如细节图10）之后 到 规格1（主规格）之前
        # 有单选/多选的保持原有选项；件数-数量、建议零售价货币、首次期望上架日期 特殊处理；其余可绑飞书或选「自己填写」
        # ================================================================
        st.markdown("---")
        st.header("🔗 4. 飞书字段绑定配置")
        st.caption("区间：最后一个细节图（如 细节图10）之后 → 规格1（主规格）之前。有单选/多选保持原样；其余可绑定飞书或选择自己填写。")

        # ── 确定范围：最后一个细节图（细节图10 等）~ 规格1（主规格）────
        feishu_start = 0
        feishu_end = len(all_cols)
        for i, c in enumerate(all_cols):
            if "细节图" in c:
                feishu_start = i + 1
        for i, c in enumerate(all_cols):
            if ("规格1" in c or "主规格" in c) and i > feishu_start:
                feishu_end = i
                break

        feishu_zone_cols = []
        for i in range(feishu_start, feishu_end):
            col_name = all_cols[i]
            cfg = mapping.get(col_name, {})
            cfg_type = cfg.get("type", "ignore")
            if cfg_type == "ignore":
                continue
            # 跳过图片类列（按类型或列名）
            if cfg_type == "image" or col_name.rstrip(')').rstrip('0123456789').endswith("图"):
                continue
            feishu_zone_cols.append(col_name)

        feishu_fields_list = st.session_state.get('feishu_fields', [])
        is_feishu_ready = len(feishu_fields_list) > 0

        if not feishu_zone_cols:
            st.info("ℹ️ 当前范围内没有需要绑定的列。")
        else:
            opts_dict = st.session_state.get('options_dict', {})
            dv_dict = st.session_state.get('dv_options_dict', {})
            feishu_bind_cols = []
            other_cols = []
            for col_name in feishu_zone_cols:
                cfg = mapping.get(col_name, {})
                ct = cfg.get("type", "")
                # 件数-数量、建议零售价货币、首次期望上架日期：强制走固定属性，不参与飞书绑定
                if "件数" in col_name and "数量" in col_name:
                    other_cols.append(col_name)
                    if ct not in ("fixed_single", "fixed_multi", "fixed_text"):
                        mapping.setdefault(col_name, {})["type"] = "fixed_text"
                elif "建议零售价货币" in col_name:
                    other_cols.append(col_name)
                    if ct not in ("fixed_single", "fixed_multi", "fixed_text"):
                        mapping.setdefault(col_name, {})["type"] = "fixed_text"
                elif "首次期望上架日期" in col_name:
                    other_cols.append(col_name)
                    if ct not in ("fixed_single", "fixed_multi", "fixed_text"):
                        mapping.setdefault(col_name, {})["type"] = "fixed_text"
                # Excel 中该列为下拉单选（有数据验证/选项）→ 按映射走固定属性单选，不绑飞书
                elif (col_name in opts_dict and opts_dict.get(col_name)) or (col_name in dv_dict and dv_dict.get(col_name)):
                    other_cols.append(col_name)
                    if ct not in ("fixed_single", "fixed_multi", "fixed_text"):
                        mapping.setdefault(col_name, {})["type"] = "fixed_single"
                elif ct in ("fixed_single", "fixed_multi", "fixed_text"):
                    other_cols.append(col_name)
                else:
                    feishu_bind_cols.append(col_name)

            # ── 单选/多选/固定文本列：保持原有组件 ────────────────
            if other_cols:
                st.caption(f"📦 固定属性 ({len(other_cols)} 项) — 保持原有单选/多选形式")
                for col_name in other_cols:
                    # 上架方式：渲染并保存到 mapping，供「首次期望上架日期」判断
                    if "上架方式" in col_name and "首次" not in col_name:
                        val = _render_attr_widget(col_name, "zone_attr")
                        if val is not None and not str(val).startswith("__FEISHU__"):
                            mapping.setdefault(col_name, {})["value"] = str(val)
                    elif "件数" in col_name and "数量" in col_name:
                        # 件数-数量：数量下拉（1-100），并持久化到 mapping
                        QTY_OPTIONS = [str(i) for i in range(1, 101)]
                        cfg = mapping.get(col_name, {})
                        prev_val = cfg.get("value", "") or cfg.get("feishu_key", "")
                        sel_idx = QTY_OPTIONS.index(prev_val) if prev_val in QTY_OPTIONS else 0
                        qty_val = st.selectbox(
                            _display_name(col_name),
                            options=QTY_OPTIONS,
                            index=sel_idx,
                            key=f"zone_qty_{col_name}"
                        )
                        mapping[col_name]["value"] = qty_val
                        mapping[col_name]["type"] = "fixed_text"
                        mapping[col_name]["feishu_key"] = ""
                    elif "建议零售价货币" in col_name:
                        # 建议零售价货币：货币下拉，默认 CNY，并持久化到 mapping
                        CURRENCY_OPTIONS = ["USD", "CNY", "EUR", "SAR", "AED", "CAD", "MXN", "HKD", "VND", "THB", "GBP", "INR", "BRL", "TRY", "NZD"]
                        cfg = mapping.get(col_name, {})
                        prev_val = cfg.get("value", "") or cfg.get("feishu_key", "")
                        sel_idx = CURRENCY_OPTIONS.index(prev_val) if prev_val in CURRENCY_OPTIONS else CURRENCY_OPTIONS.index("CNY")
                        currency_val = st.selectbox(
                            _display_name(col_name),
                            options=CURRENCY_OPTIONS,
                            index=sel_idx,
                            key=f"zone_currency_{col_name}"
                        )
                        mapping[col_name]["value"] = currency_val
                        mapping[col_name]["type"] = "fixed_text"
                        mapping[col_name]["feishu_key"] = ""
                    elif "首次期望上架日期" in col_name:
                        # 首次期望上架日期：日期选择器，若上架方式=自动上架则留空
                        shangjia_val = mapping.get("上架方式", {}).get("value", "") or st.session_state.get("zone_attr_上架方式", "")
                        if shangjia_val and "自动上架" in str(shangjia_val):
                            st.caption("ℹ️ 上架方式为「自动上架」，无需选择日期")
                            mapping[col_name]["value"] = ""
                            mapping[col_name]["type"] = "fixed_text"
                            mapping[col_name]["feishu_key"] = ""
                        else:
                            cfg = mapping.get(col_name, {})
                            prev_val = cfg.get("value", "") or cfg.get("feishu_key", "")
                            default_date = None
                            if prev_val:
                                try:
                                    default_date = datetime.strptime(prev_val[:10], "%Y-%m-%d").date()
                                except (ValueError, TypeError):
                                    pass
                            date_val = st.date_input(
                                _display_name(col_name),
                                value=default_date,
                                key=f"zone_date_{col_name}"
                            )
                            mapping[col_name]["value"] = date_val.strftime("%Y-%m-%d") if date_val else ""
                            mapping[col_name]["type"] = "fixed_text"
                            mapping[col_name]["feishu_key"] = ""
                    else:
                        # 有 Excel 下拉选项的列（如重量单位）走 _render_attr_widget，按 options_dict/dv_options_dict 渲染单选
                        val = _render_attr_widget(col_name, "zone_attr")
                        if val is not None and not str(val).startswith("__FEISHU__"):
                            mapping.setdefault(col_name, {})["value"] = str(val)

            # ── 飞书绑定列（可绑飞书字段，或选「自己填写」不强制绑定）────
            if feishu_bind_cols:
                st.caption(f"🔗 飞书绑定 ({len(feishu_bind_cols)} 项) — 可选飞书字段或「自己填写」")
                if not is_feishu_ready:
                    st.warning("⚠️ 请先在顶部「飞书环境配置」区连接飞书并拉取字段列表。")
                else:
                    st.success(f"✅ 已获取 {len(feishu_fields_list)} 个飞书字段")
                    _PRICE_X15_OPT = "价格×1.5（取整）"
                    for col_name in feishu_bind_cols:
                        cfg = mapping.get(col_name, {})
                        prev_feishu_key = cfg.get("feishu_key", "")
                        prev_type = cfg.get("type", "")
                        label = _display_name(col_name)
                        # 建议零售价（非货币）额外提供「价格×1.5（取整）」快捷选项
                        _is_jy_retail = "建议零售价" in col_name and "货币" not in col_name
                        # 不强制绑定：可选飞书字段，或「不绑定」/「自己填写」（第二步手填）
                        _base_opts = ["-- 不绑定 --", "-- 自己填写 --"]
                        if _is_jy_retail:
                            _base_opts = ["-- 不绑定 --", "-- 自己填写 --", _PRICE_X15_OPT]
                        feishu_options_with_empty = _base_opts + feishu_fields_list
                        if prev_type == "price_x1.5":
                            _display_val = _PRICE_X15_OPT
                        elif prev_type == "manual" and not prev_feishu_key:
                            _display_val = "-- 自己填写 --"
                        else:
                            _display_val = prev_feishu_key if prev_feishu_key in feishu_options_with_empty else "-- 不绑定 --"
                        sel_idx = feishu_options_with_empty.index(_display_val) if _display_val in feishu_options_with_empty else 0
                        _wk = f"feishu_bind_{col_name}"
                        if _wk not in st.session_state:
                            st.session_state[_wk] = feishu_options_with_empty[sel_idx]

                        col_label, col_select = st.columns([1, 1])
                        with col_label:
                            r = req_dict.get(col_name, '')
                            badge = "🔴" if ('必填' in r and '非必填' not in r) else "⚪"
                            st.markdown(f"{badge} **{label}**")
                        with col_select:
                            selected_field = st.selectbox(
                                "飞书字段", options=feishu_options_with_empty,
                                index=sel_idx, key=_wk,
                                label_visibility="collapsed"
                            )
                            if selected_field == _PRICE_X15_OPT:
                                mapping[col_name]["feishu_key"] = ""
                                mapping[col_name]["type"] = "price_x1.5"
                            elif selected_field not in ("-- 不绑定 --", "-- 自己填写 --"):
                                mapping[col_name]["feishu_key"] = selected_field
                                mapping[col_name]["type"] = "feishu"
                            else:
                                mapping[col_name]["feishu_key"] = ""
                                mapping[col_name]["type"] = "manual"

        # ================================================================
        # ██ 其他可选固定信息（去重后沉底） ██
        # ================================================================
        # 收集规格区和飞书区已出现的列名
        _shown_cols = set()
        # 规格区范围
        for i in range(range_start, range_end):
            _shown_cols.add(all_cols[i])
        # 飞书区范围
        for i in range(feishu_start, feishu_end):
            _shown_cols.add(all_cols[i])

        _other_cols = [c for c in st.session_state.excel_columns
                       if c not in _shown_cols and not c.startswith('_')]

        _configured_cols = []
        _unconfigured_cols = []
        _ACTIVE_TYPES = {"feishu", "fixed_single", "fixed_multi", "fixed_text", "manual"}
        for c in _other_cols:
            cfg = st.session_state.config_mapping.get(c, {})
            cfg_type = cfg.get("type", "ignore")
            # 飞书绑定 / 固定值 / 手动填写 → 归入已配置（多语言列除外）
            if cfg_type in _ACTIVE_TYPES and "多语言" not in c:
                _configured_cols.append(c)
            else:
                _unconfigured_cols.append(c)

        st.markdown("---")

        # ── ✅ 可能也需要的固定配置（直接显示） ─────────────────────
        if _configured_cols:
            with st.expander(f"📌 可能也需要的固定配置 — 共 {len(_configured_cols)} 列", expanded=True):
                st.caption("这些列可能需要设定固定值或飞书绑定，请根据实际需求配置。")
                for idx, col_name in enumerate(_configured_cols):
                    _render_col_config(col_name, idx + 9000, st.session_state.config_mapping)

        # ── 🛠️ 未配置的列（折叠） ──────────────────────────────────
        with st.expander(f"🛠️ 其他信息（如果不固定则无需选择）— 共 {len(_unconfigured_cols)} 列", expanded=False):
            st.caption('这些列默认为"无需处理/忽略"，如果某列需要固定填写，请手动切换类型并填入值。')
            for idx, col_name in enumerate(_unconfigured_cols):
                _render_col_config(col_name, idx + 19000, st.session_state.config_mapping)

        # ================================================================
        # ██ 保存模板（收集全量配置） ██
        # ================================================================
        st.markdown("---")
        st.header("💾 5. 保存模板")
        st.caption("保存以上所有区域的配置（规格属性 + 飞书绑定 + 其他固定信息），供第二步填写时使用")

        col_save_name, col_save_btn = st.columns([3, 1])
        with col_save_name:
            default_save_name = st.session_state.loaded_template_name or "shein_template_config.json"
            save_filename = st.text_input("模板文件名称 (需以 .json 结尾)", value=default_save_name)
            if not save_filename.endswith('.json'):
                save_filename += '.json'

        with col_save_btn:
            st.write("")
            st.write("")
            if st.button("💾 保存配置为模板", type="primary", use_container_width=True):
                # 收集所有非 ignore 的列配置（飞书凭证不写入模板，统一由后端 Secrets 配置，避免泄露）
                template_data = {}
                full_mapping = st.session_state.config_mapping

                # 遍历所有列，收集非 ignore 的配置（含 value、feishu_key 等，下次加载可直接反填）
                for col_name, col_cfg in full_mapping.items():
                    if col_name.startswith('_'):
                        continue
                    if col_cfg.get("type", "ignore") != "ignore":
                        template_data[col_name] = dict(col_cfg)

                # 保存内部选项字典给第二步使用
                if "options_dict" in st.session_state:
                    template_data["_options_dict"] = st.session_state.options_dict
                if "dv_options_dict" in st.session_state:
                    template_data["_dv_options_dict"] = st.session_state.dv_options_dict

                if not template_data:
                    st.warning("⚠️ 没有检测到有效配置，无法保存。")
                else:
                    try:
                        save_path = CONFIG_DIR / save_filename
                        with open(save_path, 'w', encoding='utf-8') as f:
                            json.dump(template_data, f, ensure_ascii=False, indent=2)
                        # 同时保存当前上传的 Excel，第二步加载该模板时可自动带出，无需再上传
                        excel_bytes = st.session_state.get('uploaded_excel_bytes')
                        if excel_bytes:
                            xlsx_path = CONFIG_DIR / f"{save_path.stem}.xlsx"
                            with open(xlsx_path, 'wb') as f:
                                f.write(excel_bytes)
                        saved_count = len([k for k in template_data if not k.startswith('_')])
                        msg = f"✅ 模板已保存！共 {saved_count} 个字段配置 → `{save_path}`。"
                        if excel_bytes:
                            msg += " 已同时保存 Excel 模板，第二步加载后可直接生成，无需再上传。"
                        else:
                            msg += " 下次加载此模板将直接反填已选配置。"
                        st.success(msg)
                        st.session_state.loaded_template_name = save_filename
                    except Exception as e:
                        st.error(f"❌ 保存失败: {str(e)}")

    else:
        if not uploaded_file:
            st.info("👈 还没有上传模板文件，请先上传您的 Excel 模板！")



# ================================================================
# ██ 第二步：根据货号拉取飞书数据并生成 Excel ██
# ================================================================
elif st.session_state.current_step == "step2":
    st.title("🚀 第二步：生成上传文件")
    st.markdown("加载第一步保存的模板，输入货号，从飞书拉取数据，生成最终 Excel 上传文件。")

    # ── 1. 加载模板 ──────────────────────────────────────────────
    st.subheader("📦 1. 选择模板")
    _s2_templates = get_saved_templates()
    if not _s2_templates:
        st.warning("⚠️ 还没有保存任何模板，请先完成第一步。")
        st.stop()

    _s2_col1, _s2_col2 = st.columns([3, 1])
    with _s2_col1:
        # 如果第一步选择了模板，则作为默认选中项，甚至可以自动加载
        default_tpl = st.session_state.get('loaded_template_name', "")
        
        # 尝试自动加载（如果进入第二步尚未加载任何模板，且第一步有指定模板）
        if default_tpl and default_tpl in _s2_templates and not st.session_state.get('s2_template_name'):
            try:
                _auto_s2_data, _auto_excel_bytes = load_template_with_excel(default_tpl)
                if _auto_s2_data:
                    st.session_state.s2_template = _auto_s2_data
                    st.session_state.s2_template_name = default_tpl
                    if _auto_excel_bytes:
                        st.session_state['uploaded_excel_bytes'] = _auto_excel_bytes
                        _cols, _req, _rules = _parse_excel_columns_from_bytes(_auto_excel_bytes)
                        if _cols:
                            st.session_state.excel_columns = _cols
                            st.session_state.excel_required_dict = _req
                            st.session_state.excel_rules_dict = _rules
            except Exception:
                pass

        current_sel = st.session_state.get('s2_template_name', "")
        try:
            sel_idx = _s2_templates.index(current_sel) if current_sel in _s2_templates else 0
        except ValueError:
            sel_idx = 0

        _s2_sel = st.selectbox("选择配置模板", options=_s2_templates, index=sel_idx, key="s2_template_sel")
    with _s2_col2:
        st.write("")
        st.write("")
        _s2_load_btn = st.button("📂 加载模板", type="primary", use_container_width=True)

    if _s2_load_btn:
        try:
            _s2_data, _s2_excel_bytes = load_template_with_excel(_s2_sel)
            if _s2_data:
                st.session_state.s2_template = _s2_data
                st.session_state.s2_template_name = _s2_sel
                # 若该模板曾保存过 Excel，自动恢复，无需第一步再上传
                if _s2_excel_bytes:
                    st.session_state['uploaded_excel_bytes'] = _s2_excel_bytes
                    _cols, _req, _rules = _parse_excel_columns_from_bytes(_s2_excel_bytes)
                    if _cols:
                        st.session_state.excel_columns = _cols
                        st.session_state.excel_required_dict = _req
                        st.session_state.excel_rules_dict = _rules
                    st.success(f"✅ 模板 [{_s2_sel}] 加载成功！已恢复绑定的 Excel 模板，可直接生成文件。")
                else:
                    st.success(f"✅ 模板 [{_s2_sel}] 加载成功！")
                st.rerun()
            else:
                st.error("模板为空或不存在")
        except Exception as e:
            st.error(f"加载失败: {e}")

    if 's2_template' not in st.session_state or not st.session_state.s2_template:
        st.info("👆 请先选择并加载模板。")
        st.stop()

    _tpl = st.session_state.s2_template
    _tpl_name = st.session_state.get('s2_template_name', '')
    st.caption(f"当前模板: **{_tpl_name}** — 共 {len([k for k in _tpl if not k.startswith('_')])} 个字段配置")

    # ── 2. 飞书凭证：Secrets → 模板内 → 侧栏缓存/本地文件 → config_mapping（侧栏填写并测试连接后会有）────
    _fs_auth = (
        _get_feishu_from_secrets()
        or _tpl.get("_feishu_auth")
        or _load_feishu_cache()
        or st.session_state.get("config_mapping", {}).get("_feishu_auth")
        or {}
    )
    _has_auth = all(_fs_auth.get(k) for k in ["app_id", "app_secret", "app_token", "table_id"])

    if not _has_auth:
        st.error("⚠️ 未配置飞书凭证：请在侧栏填写并测试连接，或由管理员在部署后台配置 Secrets。")
        st.stop()

    # 飞书凭证不在前台展示，避免暴露 API 密钥

    # --- 状态缓存初始化 ---
    if 's2_persisted_values' not in st.session_state:
        st.session_state.s2_persisted_values = {}

    def _sync_s2_val(key_name, field_name):
        val = st.session_state[key_name]
        if isinstance(val, list):
            val = "*".join(val)
        if val in ("-- 请选择 --", "-- 不填 --"):
            val = ""
        st.session_state.s2_persisted_values[field_name] = val

    def _add_extra_variant_sku():
        c = st.session_state.get("s2_extra_sku_count", 0)
        st.session_state["s2_extra_sku_count"] = min(29, c + 1)
        if "s2_persisted_values" in st.session_state:
            st.session_state.s2_persisted_values["s2_extra_sku_count"] = st.session_state["s2_extra_sku_count"]

    # ── 上传模式切换 ────────────────────────────────────────────────────
    st.markdown("---")
    _upload_mode = st.radio(
        "请选择上传模式",
        ["单品上传", "复色上传 (多SKU同款)"],
        horizontal=True,
        key="s2_upload_mode"
    )
    _is_multi_variant = _upload_mode == "复色上传 (多SKU同款)"

    _gen_mode = st.radio(
        "生成方式",
        ["单条生成", "批量生成"],
        horizontal=True,
        key="s2_gen_mode",
        help="批量生成时在下方表格中填写多行（供方货号、卖家SKU、标题、五点描述等），每行一条或一组复色。"
    )
    _use_batch = _gen_mode == "批量生成"

    # ── 3. 核心输入：检索与货号 ──────────────────────────────────────
    st.subheader("🔑 2. 输入核心检索信息")

    # 获取飞书字段列表（用于选择匹配字段）
    _s2_feishu_fields = st.session_state.get('feishu_fields', [])

    st.markdown("**(1) 飞书检索凭证**")
    _field_options = ["货号", "卖家SKU"] + [f for f in _s2_feishu_fields if f not in ("货号", "卖家SKU")]
    prev_match = st.session_state.s2_persisted_values.get("s2_match_field", "货号")
    sel_idx = _field_options.index(prev_match) if prev_match in _field_options else 0
    _wk_match = "s2_match_field"
    _match_field = st.selectbox(
        "飞书中对应的检索列",
        options=_field_options, index=sel_idx,
        help="单条生成：下方输入的检索值将按此列匹配飞书。批量生成：表格中每行的「供方货号」或「卖家SKU」等将按此列匹配飞书，只需选对应字段即可。",
        key=_wk_match, on_change=_sync_s2_val, args=(_wk_match, "s2_match_field")
    )
    if not _use_batch:
        _col_sku1, _col_sku2 = st.columns([1, 2])
        with _col_sku1:
            st.caption("检索值（单条生成必填）")
        with _col_sku2:
            prev_sku = st.session_state.s2_persisted_values.get("s2_sku_input", "")
            _wk_sku = "s2_sku_input"
            _input_sku = st.text_input(
                "📝 请输入检索值 (必填，将作为飞书检索的唯一凭证)", value=prev_sku,
                placeholder="根据上方检索列填写，如选货号则填货号、选卖家SKU则填卖家SKU",
                key=_wk_sku, on_change=_sync_s2_val, args=(_wk_sku, "s2_sku_input"),
                label_visibility="collapsed"
            )
            if _is_multi_variant:
                if "s2_extra_sku_count" not in st.session_state and "s2_persisted_values" in st.session_state:
                    st.session_state["s2_extra_sku_count"] = st.session_state.s2_persisted_values.get("s2_extra_sku_count", 0)
                _extra_count = st.session_state.get("s2_extra_sku_count", 0)
                st.button("➕ 添加复色卖家SKU", key="s2_btn_add_extra_sku", on_click=_add_extra_variant_sku, help="每点击一次增加一个复色变体（该 SKU 与上方主 SKU 同款不同色/规格）")
                for i in range(_extra_count):
                    _prev_extra = st.session_state.s2_persisted_values.get(f"s2_extra_sku_{i}", "")
                    st.text_input(
                        f"复色卖家SKU #{i+2}",
                        value=st.session_state.get(f"s2_extra_sku_{i}", _prev_extra),
                        placeholder=f"第 {i+2} 个变体的卖家SKU",
                        key=f"s2_extra_sku_{i}",
                        on_change=_sync_s2_val,
                        args=(f"s2_extra_sku_{i}", f"s2_extra_sku_{i}")
                    )
        st.markdown("**(2) 供方货号（复色时所有行共用同一个供方货号）**")
        prev_huohao = st.session_state.s2_persisted_values.get("s2_huohao_input", "")
        _wk_huohao = "s2_huohao_input"
        _input_huohao = st.text_input(
            "📝 请输入 供方货号 (纯手填，直接写入表格；复色时每行都填此同一货号)", value=prev_huohao,
            placeholder="例如: 456（复色时 321、123 两个 SKU 都填供方货号 456）",
            key=_wk_huohao, on_change=_sync_s2_val, args=(_wk_huohao, "s2_huohao_input")
        )
    else:
        st.caption("批量生成时，表格中每行已有多个供方货号/卖家SKU，只需选择上方「飞书中对应的检索列」与表格中哪一列对应即可（如选卖家SKU 则按表格的卖家SKU 列去飞书匹配）。")
        _input_sku = ""
        _input_huohao = ""

    # ── 复色模式：变体数量与 SKU 列表（规格2内容改在「核心动态录入区」内按货号填写）────
    _excel_cols_early = st.session_state.get("excel_columns", [])
    _opts_early = st.session_state.get("options_dict", _tpl.get("_options_dict", {}))
    _dv_opts_early = st.session_state.get("dv_options_dict", _tpl.get("_dv_options_dict", {}))
    _spec2_opts = []
    for _c in _excel_cols_early:
        if _c and "规格2内容" in _c and _c in _dv_opts_early and _dv_opts_early[_c]:
            _spec2_opts = list(_dv_opts_early[_c])
            break
    if not _spec2_opts:
        for _dk, _vals in _opts_early.items():
            if "规格2内容" in _dk or (_dk in _opts_early and hasattr(_vals, "__iter__") and not isinstance(_vals, str)):
                _spec2_opts = list(_vals) if hasattr(_vals, "__iter__") else []
                break

    if _is_multi_variant:
        _variant_count = 1 + st.session_state.get("s2_extra_sku_count", 0)
        _variant_skus = [_input_sku] + [st.session_state.get(f"s2_extra_sku_{i}", st.session_state.s2_persisted_values.get(f"s2_extra_sku_{i}", "")) for i in range(st.session_state.get("s2_extra_sku_count", 0))]
    else:
        _variant_count = 1
        _variant_skus = [_input_sku]

    # ── 4. 手动填写与补充配置区 ─────────────────────────────────────────────
    st.markdown("---")

    _excel_cols = st.session_state.get("excel_columns", [])
    _col_order = {c: i for i, c in enumerate(_excel_cols)}

    # 核心动态录入区列（与下方「3. 核心动态录入区」一致）：zone_a（卖家SKU→首图）+ zone_b（规格1～规格3内容）
    _sku_idx = None
    for _i, _c in enumerate(_excel_cols):
        if "卖家SKU" in _c or "供方货号" in _c:
            _sku_idx = _i
            break
    _img_idx = None
    for _i, _c in enumerate(_excel_cols):
        if "首图" in _c or "主图" in _c:
            _img_idx = _i
            break
    _zone_a = []
    if _sku_idx is not None and _img_idx is not None and _img_idx > _sku_idx:
        for _i in range(_sku_idx + 1, _img_idx):
            _c = _excel_cols[_i]
            if "多语言商品名称" in _c or "多语言商品描述" in _c:
                continue
            if not _c.startswith("_"):
                _zone_a.append(_c)
    _spec_keywords = ["规格1", "规格1内容", "主规格", "规格2", "规格2内容", "规格3", "规格3内容"]
    _zone_b = [_c for _c in _excel_cols if not _c.startswith("_") and any(_kw in _c for _kw in _spec_keywords)]
    _required_fill_cols = []
    _dynamic_set = set()
    for _c in _zone_a + _zone_b:
        if _c not in _dynamic_set:
            _dynamic_set.add(_c)
            _required_fill_cols.append(_c)
    _required_fill_cols.sort(key=lambda c: _col_order.get(c, 9999))
    for _col in _required_fill_cols:
        if _col not in _tpl:
            _tpl[_col] = {"type": "fixed_text", "value": ""}
        elif _tpl[_col].get("type") == "ignore":
            _tpl[_col] = {**_tpl[_col], "type": "fixed_text"}

    _huohao_col = next((c for c in _excel_cols if c and ("供方货号" in c or "货号" in c)), "供方货号")
    _sku_col = next((c for c in _excel_cols if c and ("卖家SKU" in c or "SKU" in c.upper())), "卖家SKU")
    # 批量表列 = 供方货号 + 卖家SKU + 核心动态录入区全部列（与下方「核心动态录入区」一致，每行可填该货号的标题、五点描述、品牌、规格等）
    _batch_df_columns = [_huohao_col, _sku_col] + [c for c in _required_fill_cols if c not in (_huohao_col, _sku_col)]
    # 规格123内容在表格内增加「手填」列，实现同行内既可下拉单选又可自己填写
    _spec_content_cols = [c for c in _batch_df_columns if "规格1内容" in c or "规格2内容" in c or "规格3内容" in c]
    _batch_df_columns_table = []
    for c in _batch_df_columns:
        _batch_df_columns_table.append(c)
        if c in _spec_content_cols:
            _batch_df_columns_table.append(c + "_手填")
    # 左侧自动编号列，便于与下方案单「第几行」对应；只读，粘贴不会覆盖
    _batch_df_columns_table = ["序号"] + _batch_df_columns_table

    # 批量录入表（仅批量生成时展示）：列与下方「核心动态录入区」一致，每行可填该货号的标题、五点描述、品牌、规格等
    # 用 fragment 隔离：表格内下拉第一次选择时只重跑本块，不触发整页加载，避免选择被整页 rerun 覆盖而消失
    if _use_batch:
        st.markdown("---")
        st.subheader("📋 批量录入表（与核心动态录入区一致）")
        st.caption("左侧「序号」列自动编号（1,2,3…），与下方批量录入「当前编辑第几行」对应，只读不可编辑。从 Excel 粘贴时请从「货号」列开始粘贴。规格1/2/3内容：每项两列（下拉+手填），生成时优先取手填列。")
        _batch_key = "s2_batch_df"
        _cols_key = "s2_batch_df_columns"
        _prev_cols = st.session_state.get(_cols_key)
        _cols_unchanged = _prev_cols is not None and len(_prev_cols) == len(_batch_df_columns_table) and all(a == b for a, b in zip(_prev_cols, _batch_df_columns_table))
        if _batch_key not in st.session_state or not _cols_unchanged:
            _init_df = pd.DataFrame({c: ([1] if c == "序号" else [""]) for c in _batch_df_columns_table})
            st.session_state[_batch_key] = _init_df
            st.session_state[_cols_key] = list(_batch_df_columns_table)
            st.session_state["s2_edited_batch_df"] = _init_df.copy()
            st.session_state["s2_batch_df_stable"] = _init_df.copy()
            if "s2_batch_editor" in st.session_state:
                del st.session_state["s2_batch_editor"]

        def _sync_batch_editor_to_session():
            """（已弃用 on_change，改为渲染后返回值检测）保留此函数供 _sync_form_row 等内部引用。"""
            pass

        # s2_batch_df_stable：传给 data_editor 的 data 参数。每次 on_change 后与编辑结果对齐，保证 data 参数不会因外部逻辑变化而触发 widget 重新初始化
        _batch_df = st.session_state.get(_batch_key)
        if _batch_df is not None and isinstance(_batch_df, pd.DataFrame):
            # 初始化 stable base（只在没有时才建，后续由 on_change 维护）
            if "s2_batch_df_stable" not in st.session_state:
                _init_stable = st.session_state.get("s2_edited_batch_df") or _batch_df
                st.session_state["s2_batch_df_stable"] = _init_stable.copy()
            _batch_df_to_show = st.session_state["s2_batch_df_stable"].copy()
            if list(_batch_df_to_show.columns) != list(_batch_df_columns_table):
                _batch_df_to_show = _batch_df
                st.session_state["s2_batch_df_stable"] = _batch_df.copy()
            _batch_df_to_show = _batch_df_to_show.reset_index(drop=True)
            _batch_df_to_show["序号"] = list(range(1, len(_batch_df_to_show) + 1))
            _batch_column_config = {}
            for col_name in _batch_df_columns_table:
                if col_name == "序号":
                    _batch_column_config[col_name] = st.column_config.NumberColumn(
                        label="序号",
                        format="%d",
                        disabled=True,
                        help="自动编号，与下方「当前编辑第几行」对应",
                    )
                    continue
                if col_name.endswith("_手填"):
                    # 规格X内容_手填：同行内既可下拉又可自己填，手填列优先
                    _batch_column_config[col_name] = st.column_config.TextColumn(
                        label=_display_name(col_name.replace("_手填", "")) + "(手填)",
                        default="",
                        help="直接填写时优先于左侧下拉",
                    )
                    continue
                cfg_type = _tpl.get(col_name, {}).get("type", "ignore")
                opts = []
                if col_name in _dv_opts_early and _dv_opts_early[col_name]:
                    opts = list(_dv_opts_early[col_name])
                else:
                    for dk, dv_vals in _opts_early.items():
                        if dk in col_name or col_name in dk:
                            opts = list(dv_vals) if hasattr(dv_vals, "__iter__") and not isinstance(dv_vals, str) else []
                            break
                if cfg_type == "fixed_single" and opts:
                    is_spec_content = col_name in _spec_content_cols
                    options_list = ["", "-- 请选择 --", "-- 不填 --"]
                    if is_spec_content:
                        options_list.append("-- 自己填写 --")
                    options_list.extend(opts)
                    if col_name in _batch_df_to_show.columns:
                        for _v in _batch_df_to_show[col_name].dropna().astype(str).unique():
                            _v = (_v or "").strip()
                            if _v and _v not in options_list and _v.lower() != "nan":
                                options_list.append(_v)
                    _col_kw = {"label": _display_name(col_name), "options": options_list, "default": ""}
                    if is_spec_content:
                        _col_kw["help"] = "可下拉选择；或在本行右侧「(手填)」列直接填写"
                    _batch_column_config[col_name] = st.column_config.SelectboxColumn(**_col_kw)
            # 为货号/卖家SKU列在表头显示「(供方货号)」/「(实际货号)」提示
            for _hint_col in [_huohao_col, _sku_col]:
                if _hint_col in _batch_df_columns_table:
                    _batch_column_config[_hint_col] = st.column_config.TextColumn(
                        label=_display_name(_hint_col),
                        default="",
                    )
            _edited_batch_df = st.data_editor(
                _batch_df_to_show,
                key="s2_batch_editor",
                num_rows="dynamic",
                use_container_width=True,
                column_config=_batch_column_config,
                hide_index=True,
            )
            # 每次渲染后用返回值与已保存数据对比，检测表格编辑，实现表格→表单联动。
            # 关键：只更新 s2_edited_batch_df 和 version，不动 s2_batch_df_stable。
            # 若同时更新 stable，则下次渲染时 data 参数变化会触发 data_editor 重新初始化，导致第二次输入消失。
            # stable 只由表单写回（_sync_form_row）和加/删行按钮维护，保证 data 参数对 data_editor 始终稳定。
            # 在归一化前记录原始序号，用于判断是否需要强制修正（如粘贴新行导致序号为 None）
            _orig_seq = _edited_batch_df["序号"].tolist() if "序号" in _edited_batch_df.columns else []
            _ret_df = _edited_batch_df.copy()
            for _c in list(_ret_df.columns):
                if _c != "序号":
                    _ret_df[_c] = _ret_df[_c].astype(object).fillna("").astype(str).replace("nan", "")
            _ret_df = _ret_df.reset_index(drop=True)
            _ret_df["序号"] = list(range(1, len(_ret_df) + 1))
            _prev_saved = st.session_state.get("s2_edited_batch_df")
            _tbl_changed = False
            try:
                if _prev_saved is None or _prev_saved.shape != _ret_df.shape or not _prev_saved.equals(_ret_df):
                    _tbl_changed = True
            except Exception:
                _tbl_changed = True
            if _tbl_changed:
                st.session_state["s2_edited_batch_df"] = _ret_df
                st.session_state[_batch_key] = _ret_df
                # 不更新 s2_batch_df_stable，避免 data_editor data 参数变化触发重新初始化
                st.session_state["s2_batch_df_version"] = st.session_state.get("s2_batch_df_version", 0) + 1
                # 若序号列有误（如粘贴新行后序号为 None/不连续），强制更新 stable 并重置表格以显示正确序号
                _correct_seq = list(range(1, len(_ret_df) + 1))
                try:
                    _seq_wrong = [int(v) if v not in (None, "", "None", "nan") else None for v in _orig_seq] != _correct_seq
                except Exception:
                    _seq_wrong = True
                if _seq_wrong:
                    st.session_state["s2_batch_df_stable"] = _ret_df.copy()
                    if "s2_batch_editor" in st.session_state:
                        del st.session_state["s2_batch_editor"]
                    st.rerun()

    # _required_fill_cols、_dynamic_set 已在上方与批量表列一并计算，此处直接使用

    # 模板已固定的规格与属性：来自模板的固定/飞书列，且不在核心动态录入区
    _prefilled_fixed_cols = []
    for col_name, col_cfg in _tpl.items():
        if col_name.startswith("_"):
            continue
        if col_name in _dynamic_set:
            continue
        if "货号" in col_name or "卖家SKU" in col_name or "SKU" in col_name.upper() or "多语言" in col_name:
            continue
        cfg_type = col_cfg.get("type", "ignore")
        tpl_val = col_cfg.get("value", "")
        if tpl_val in ("-- 请选择 --", "-- 不填 --"):
            tpl_val = ""
        if cfg_type in ("fixed_single", "fixed_multi", "fixed_text") and tpl_val:
            _prefilled_fixed_cols.append(col_name)
        elif cfg_type == "feishu":
            _prefilled_fixed_cols.append(col_name)
    _prefilled_fixed_cols.sort(key=lambda c: _col_order.get(c, 9999))

    _s2_manual_values = {}
    # 尝试从当前会话拿，拿不到就尝试从模板里恢复
    _dv_opts = st.session_state.get('dv_options_dict', _tpl.get('_dv_options_dict', {}))
    _opts_dict = st.session_state.get('options_dict', _tpl.get('_options_dict', {}))

    def _render_s2_field(ec, is_prefilled=False, is_required=False):
        """共有函数：在第二步界面基于模板渲染并同步指定的列输入形式。is_required 为 True 时在标签旁标注必填。"""
        cfg_type = _tpl.get(ec, {}).get("type", "ignore")
        tpl_val = _tpl.get(ec, {}).get("value", "")

        _req_badge = " 🔴必填" if is_required else ""
        _wk = f"s2_edit_{ec}"
        # 对于有默认值的固定列，第一次加载时也应展示配置里的初始值
        prev_val = st.session_state.s2_persisted_values.get(ec, tpl_val if is_prefilled else "")
        if prev_val in ("-- 请选择 --", "-- 不填 --"):
            prev_val = ""

        # 获取该列的下拉选项
        opts = []
        if ec in _dv_opts and _dv_opts[ec]:
            opts = list(_dv_opts[ec])
        else:
            for dk, dv_vals in _opts_dict.items():
                if dk in ec or ec in dk:
                    opts = list(dv_vals)
                    break

        if cfg_type == "fixed_single" and opts:
            # 规格X内容：单选与手填合并为一行（左下拉右输入）
            _spec_content = "规格1内容" in ec or "规格2内容" in ec or "规格3内容" in ec
            _manual_fill = "-- 自己填写 --"
            full_opts = ["-- 请选择 --", "-- 不填 --", _manual_fill] + opts if _spec_content else ["-- 请选择 --", "-- 不填 --"] + opts
            _custom_default = (prev_val or tpl_val) if ((prev_val or tpl_val) and (prev_val or tpl_val) not in full_opts) else ""
            sel_idx = full_opts.index(prev_val) if prev_val in full_opts else (full_opts.index(tpl_val) if tpl_val in full_opts else (full_opts.index(_manual_fill) if _custom_default else 0))
            if _spec_content:
                st.caption(_display_name(ec) + _req_badge + " 🔘单选/手填（选预设或右侧直接输入）")
                _sc0, _sc1 = st.columns([1, 1.2])
                with _sc0:
                    val = st.selectbox("预设", options=full_opts, index=sel_idx, key=_wk, label_visibility="collapsed", on_change=_sync_s2_val, args=(_wk, ec))
                with _sc1:
                    _custom_val = st.text_input("手填", value=_custom_default, key=_wk + "_custom", placeholder="选「自己填写」或在此输入", label_visibility="collapsed", on_change=_sync_s2_val, args=(_wk + "_custom", ec))
                if _custom_val and str(_custom_val).strip():
                    _s2_manual_values[ec] = str(_custom_val).strip()
                elif val not in ("-- 请选择 --", "-- 不填 --", _manual_fill):
                    _s2_manual_values[ec] = val
                else:
                    _s2_manual_values[ec] = ""
            else:
                val = st.selectbox(
                    _display_name(ec) + _req_badge + " 🔘单选",
                    options=full_opts, index=sel_idx,
                    key=_wk, on_change=_sync_s2_val, args=(_wk, ec)
                )
                _s2_manual_values[ec] = val if val not in ("-- 请选择 --", "-- 不填 --") else ""
        elif cfg_type == "fixed_multi" and opts:
            # 多选：直接使用下拉
            prev_list = [v.strip() for v in str(prev_val).split("*")] if prev_val else ([v.strip() for v in str(tpl_val).split("*")] if tpl_val else [])
            valid_prev = [v for v in prev_list if v in opts]
            selected = st.multiselect(
                _display_name(ec) + _req_badge + " ☑️多选",
                options=opts, default=valid_prev,
                key=_wk, on_change=_sync_s2_val, args=(_wk, ec)
            )
            _s2_manual_values[ec] = "*".join(selected)
        elif cfg_type == "feishu":
            feishu_key = _tpl.get(ec, {}).get("feishu_key", "")
            _s2_manual_values[ec] = st.text_input(
                _display_name(ec) + _req_badge + " 🔗飞书动态匹配", value=f"[自动拉取]: {feishu_key}",
                disabled=True, key=_wk
            )
        else:
            # 获取不到选项，或普通文本、manual时降级为输入框
            label_suffix = " ✍️手填" if cfg_type == "manual" else (" 📌固定文本" if cfg_type == "fixed_text" else "")
            _s2_manual_values[ec] = st.text_input(
                _display_name(ec) + _req_badge + label_suffix, value=prev_val,
                key=_wk, on_change=_sync_s2_val, args=(_wk, ec)
            )

    def _sync_form_row(row_idx, col_name):
        """表单字段变更时写回批量表 DataFrame，实现表格与表单联动。"""
        _ver = st.session_state.get("s2_batch_df_version", 0)
        key = f"s2_form_r{row_idx}_{col_name}_v{_ver}"
        df = st.session_state.get("s2_edited_batch_df")
        if df is None or not isinstance(df, pd.DataFrame):
            df = st.session_state.get("s2_batch_df")
        if df is None or not isinstance(df, pd.DataFrame) or row_idx < 1 or row_idx > len(df):
            return
        val = st.session_state.get(key)
        if isinstance(val, list):
            val = "*".join(val)
        if val in ("-- 请选择 --", "-- 不填 --", "-- 自己填写 --"):
            val = ""
        val = val if val else ""
        if col_name not in df.columns:
            return
        df = df.copy()
        df.at[row_idx - 1, col_name] = val
        st.session_state["s2_edited_batch_df"] = df
        st.session_state["s2_batch_df"] = df
        # 同步 stable base，使表格重新初始化时拿到最新数据
        st.session_state["s2_batch_df_stable"] = df.copy()
        # 递增 version，使表格 widget 重建后、form 下次渲染时也能读到新值（双向联动）
        st.session_state["s2_batch_df_version"] = st.session_state.get("s2_batch_df_version", 0) + 1
        if "s2_batch_editor" in st.session_state:
            del st.session_state["s2_batch_editor"]

    def _sync_form_row_custom(row_idx, col_name):
        """表单「自己填写」框变更时写回批量表 DataFrame，实现表格与表单联动。"""
        _ver = st.session_state.get("s2_batch_df_version", 0)
        key = f"s2_form_r{row_idx}_{col_name}_custom_v{_ver}"
        df = st.session_state.get("s2_edited_batch_df")
        if df is None or not isinstance(df, pd.DataFrame):
            df = st.session_state.get("s2_batch_df")
        if df is None or not isinstance(df, pd.DataFrame) or row_idx < 1 or row_idx > len(df):
            return
        val = (st.session_state.get(key) or "").strip()
        if col_name not in df.columns:
            return
        df = df.copy()
        df.at[row_idx - 1, col_name] = val
        st.session_state["s2_edited_batch_df"] = df
        st.session_state["s2_batch_df"] = df
        # 同步 stable base，使表格重新初始化时拿到最新数据
        st.session_state["s2_batch_df_stable"] = df.copy()
        # 递增 version，使表格 widget 重建后、form 下次渲染时也能读到新值（双向联动）
        st.session_state["s2_batch_df_version"] = st.session_state.get("s2_batch_df_version", 0) + 1
        if "s2_batch_editor" in st.session_state:
            del st.session_state["s2_batch_editor"]

    def _render_s2_field_for_row(ec, row_idx, row_dict, is_required=False, _key_version=None):
        """与核心动态录入区同款控件，按行 key 隔离，回写 row_dict。_key_version 用于表格→表单同步时使 key 随表格数据版本变化。"""
        cfg_type = _tpl.get(ec, {}).get("type", "ignore")
        tpl_val = _tpl.get(ec, {}).get("value", "")
        _req_badge = " 🔴必填" if is_required else ""
        _wk = f"s2_form_r{row_idx}_{ec}_v{_key_version}" if _key_version is not None else f"s2_form_r{row_idx}_{ec}"
        prev_val = row_dict.get(ec, "")
        if prev_val in ("-- 请选择 --", "-- 不填 --"):
            prev_val = ""
        opts = []
        if ec in _dv_opts and _dv_opts[ec]:
            opts = list(_dv_opts[ec])
        else:
            for dk, dv_vals in _opts_dict.items():
                if dk in ec or ec in dk:
                    opts = list(dv_vals)
                    break
        if cfg_type == "fixed_single" and opts:
            _spec_content = "规格1内容" in ec or "规格2内容" in ec or "规格3内容" in ec
            _manual_fill = "-- 自己填写 --"
            full_opts = ["-- 请选择 --", "-- 不填 --", _manual_fill] + opts if _spec_content else ["-- 请选择 --", "-- 不填 --"] + opts
            _custom_default = (prev_val or tpl_val) if ((prev_val or tpl_val) and (prev_val or tpl_val) not in full_opts) else ""
            sel_idx = full_opts.index(prev_val) if prev_val in full_opts else (full_opts.index(tpl_val) if tpl_val in full_opts else (full_opts.index(_manual_fill) if _custom_default else 0))
            if _spec_content:
                st.caption(_display_name(ec) + _req_badge + " 🔘单选/手填（选预设或右侧直接输入）")
                _sc0, _sc1 = st.columns([1, 1.2])
                with _sc0:
                    val = st.selectbox("预设", options=full_opts, index=sel_idx, key=_wk, label_visibility="collapsed", on_change=_sync_form_row, args=(row_idx, ec))
                with _sc1:
                    _custom_val = st.text_input("手填", value=_custom_default, key=_wk + "_custom", placeholder="选「自己填写」或在此输入", label_visibility="collapsed", on_change=_sync_form_row_custom, args=(row_idx, ec))
                if _custom_val and str(_custom_val).strip():
                    row_dict[ec] = str(_custom_val).strip()
                elif val not in ("-- 请选择 --", "-- 不填 --", _manual_fill):
                    row_dict[ec] = val
                else:
                    row_dict[ec] = ""
            else:
                val = st.selectbox(
                    _display_name(ec) + _req_badge + " 🔘单选",
                    options=full_opts, index=sel_idx,
                    key=_wk, on_change=_sync_form_row, args=(row_idx, ec)
                )
                row_dict[ec] = val if val not in ("-- 请选择 --", "-- 不填 --") else ""
        elif cfg_type == "fixed_multi" and opts:
            prev_list = [v.strip() for v in str(prev_val).split("*")] if prev_val else []
            valid_prev = [v for v in prev_list if v in opts]
            selected = st.multiselect(
                _display_name(ec) + _req_badge + " ☑️多选",
                options=opts, default=valid_prev,
                key=_wk, on_change=_sync_form_row, args=(row_idx, ec)
            )
            row_dict[ec] = "*".join(selected)
        elif cfg_type == "feishu":
            feishu_key = _tpl.get(ec, {}).get("feishu_key", "")
            st.text_input(
                _display_name(ec) + _req_badge + " 🔗飞书动态匹配", value=f"[自动拉取]: {feishu_key}",
                disabled=True, key=_wk
            )
        else:
            label_suffix = " ✍️手填" if cfg_type == "manual" else (" 📌固定文本" if cfg_type == "fixed_text" else "")
            v = st.text_input(
                _display_name(ec) + _req_badge + label_suffix, value=prev_val,
                key=_wk, on_change=_sync_form_row, args=(row_idx, ec)
            )
            row_dict[ec] = v if v else ""

    if _use_batch:
        st.markdown("---")
        st.subheader("📋 批量录入（表单式，与核心动态录入区同款控件）")
        st.caption("与上方「批量录入表」联动：在此填写或在上方表格填写，两边数据实时同步。按行展示，字段横向排列。")
        # 以批量表 DataFrame 为唯一数据源，表单只读/写该表，实现表格与表单联动
        _form_df = st.session_state.get("s2_edited_batch_df")
        if _form_df is None or not isinstance(_form_df, pd.DataFrame):
            _form_df = st.session_state.get("s2_batch_df")
        if _form_df is None or not isinstance(_form_df, pd.DataFrame) or list(_form_df.columns) != list(_batch_df_columns_table):
            _form_df = pd.DataFrame({c: [""] for c in _batch_df_columns_table})
            st.session_state["s2_edited_batch_df"] = _form_df
            st.session_state["s2_batch_df"] = _form_df
        _nr = len(_form_df)
        if _nr == 0:
            _form_df = pd.DataFrame({c: [""] for c in _batch_df_columns_table})
            st.session_state["s2_edited_batch_df"] = _form_df
            st.session_state["s2_batch_df"] = _form_df
            _nr = 1
        # 添加/删除行时通过 pending 在下一轮运行前写入行号，避免在 widget 创建后修改其 key 导致 StreamlitAPIException
        if "s2_form_pending_row_idx" in st.session_state:
            _pending = st.session_state.pop("s2_form_pending_row_idx", 1)
            st.session_state["s2_form_row_idx"] = max(1, min(int(_pending), _nr))
        _row_idx_sel = st.selectbox("当前编辑第几行", options=list(range(1, _nr + 1)), key="s2_form_row_idx", format_func=lambda x: f"第 {x} 行")
        _c_btn1, _c_btn2, _ = st.columns([1, 1, 2])
        with _c_btn1:
            if st.button("➕ 添加一行", key="s2_form_add_row"):
                _new_row = pd.DataFrame([{c: "" for c in _batch_df_columns_table}])
                _form_df = st.session_state.get("s2_edited_batch_df")
                if _form_df is None or not isinstance(_form_df, pd.DataFrame):
                    _form_df = st.session_state.get("s2_batch_df")
                if _form_df is None or not isinstance(_form_df, pd.DataFrame):
                    _form_df = pd.DataFrame({c: [""] for c in _batch_df_columns_table})
                _form_df = pd.concat([_form_df, _new_row], ignore_index=True)
                st.session_state["s2_edited_batch_df"] = _form_df
                st.session_state["s2_batch_df"] = _form_df
                st.session_state["s2_batch_df_stable"] = _form_df.copy()
                if "s2_batch_editor" in st.session_state:
                    del st.session_state["s2_batch_editor"]
                st.session_state["s2_form_pending_row_idx"] = len(_form_df)
                st.rerun()
        with _c_btn2:
            if _nr > 1 and st.button("➖ 删除当前行", key="s2_form_del_row"):
                _idx = st.session_state.get("s2_form_row_idx", 1) - 1
                _form_df = st.session_state.get("s2_edited_batch_df")
                if _form_df is None or not isinstance(_form_df, pd.DataFrame):
                    _form_df = st.session_state.get("s2_batch_df")
                if _form_df is not None and isinstance(_form_df, pd.DataFrame):
                    _form_df = _form_df.drop(_idx).reset_index(drop=True)
                    st.session_state["s2_edited_batch_df"] = _form_df
                    st.session_state["s2_batch_df"] = _form_df
                    st.session_state["s2_batch_df_stable"] = _form_df.copy()
                    if "s2_batch_editor" in st.session_state:
                        del st.session_state["s2_batch_editor"]
                    _new_nr = len(_form_df)
                    st.session_state["s2_form_pending_row_idx"] = max(1, min(st.session_state.get("s2_form_row_idx", 1), _new_nr))
                    st.rerun()
        _row_idx = min(_row_idx_sel, _nr)
        _row_dict = _form_df.iloc[_row_idx - 1].to_dict()
        for _k, _v in _row_dict.items():
            if pd.isna(_v) or _v is None:
                _row_dict[_k] = ""
            else:
                _row_dict[_k] = str(_v).strip() if isinstance(_v, str) else _v
        # 规格123内容：表格内手填列优先，表单展示合并后的值
        for _sc in _spec_content_cols:
            _hand = (_row_dict.get(_sc + "_手填") or "").strip()
            _row_dict[_sc] = _hand if _hand else (_row_dict.get(_sc) or "")
        _s2_req_dict_form = st.session_state.get("excel_required_dict", {})
        _cols_per_row = 4  # 每行 4 列，类似 Excel 横向排列，减少纵向滚动
        _batch_df_version = st.session_state.get("s2_batch_df_version", 0)  # 表格编辑后递增，使表单 widget key 变化从而用最新数据重绘
        with st.expander(f"📌 第 {_row_idx} 行：展开填写", expanded=True):
            for start in range(0, len(_batch_df_columns), _cols_per_row):
                _chunk = _batch_df_columns[start : start + _cols_per_row]
                _streamlit_cols = st.columns(len(_chunk))
                for _i, ec in enumerate(_chunk):
                    with _streamlit_cols[_i]:
                        _r = _s2_req_dict_form.get(ec, "")
                        _is_req = "必填" in _r and "非必填" not in _r
                        _render_s2_field_for_row(ec, _row_idx, _row_dict, is_required=_is_req, _key_version=_batch_df_version)

    # 单条生成时才展示核心动态录入区；批量生成时通过「批量录入表」或「表单式」按行填写，此处不展示
    if not _use_batch:
        if _required_fill_cols:
            st.subheader(f"✏️ 3. 核心动态录入区 ({len(_required_fill_cols)} 项)")
            st.caption("① 默认商品名称/描述、品牌、商品IP 等（卖家SKU→首图之间，已排除多语言列） ② 规格1～规格3内容。🔴必填 表示 Excel 中该列为必填。复色时「规格2内容」在此按每个卖家SKU 分别填写，无需重复。")
            _s2_req_dict = st.session_state.get("excel_required_dict", {})
            with st.expander("📌 展开填写", expanded=True):
                for ec in _required_fill_cols:
                    _r = _s2_req_dict.get(ec, "")
                    _is_req = "必填" in _r and "非必填" not in _r
                    if _is_multi_variant and "规格2内容" in ec:
                        st.markdown(f"**{_display_name(ec)}**（复色：每个卖家SKU 对应一行，可单选或自己填写）" + (" 🔴必填" if _is_req else ""))
                        _manual_fill = "-- 自己填写 --"
                        for i in range(_variant_count):
                            _lab = _variant_skus[i] if (i < len(_variant_skus) and _variant_skus[i]) else f"第{i+1}个SKU"
                            _c1, _c2 = st.columns([1, 2])
                            with _c1:
                                st.caption(f"卖家SKU: **{_lab}**")
                            with _c2:
                                _prev_spec = st.session_state.s2_persisted_values.get(f"s2_multi_spec2_{i}", "")
                                _prev_custom = (st.session_state.s2_persisted_values.get(f"s2_multi_spec2_{i}_custom", "") or "").strip()
                                if _spec2_opts:
                                    full_opts = ["-- 请选择 --", "-- 不填 --", _manual_fill] + list(_spec2_opts)
                                    _effective = _prev_custom if _prev_custom else (_prev_spec if _prev_spec not in ("-- 请选择 --", "-- 不填 --", _manual_fill) else "")
                                    _custom_default = _effective if (_effective and _effective not in _spec2_opts) else ""
                                    _sel_val = _prev_spec if _prev_spec in full_opts else (_manual_fill if _custom_default or _effective else "-- 请选择 --")
                                    _idx = full_opts.index(_sel_val) if _sel_val in full_opts else 0
                                    _sc0, _sc1 = st.columns([1, 1.2])
                                    with _sc0:
                                        st.selectbox(
                                            "规格2内容",
                                            options=full_opts,
                                            index=_idx,
                                            key=f"s2_multi_spec2_{i}",
                                            label_visibility="collapsed",
                                            on_change=_sync_s2_val,
                                            args=(f"s2_multi_spec2_{i}", f"s2_multi_spec2_{i}")
                                        )
                                    with _sc1:
                                        st.text_input("或直接填写", value=st.session_state.get(f"s2_multi_spec2_{i}_custom", _custom_default), key=f"s2_multi_spec2_{i}_custom", placeholder="选「自己填写」或在此输入", label_visibility="collapsed", on_change=_sync_s2_val, args=(f"s2_multi_spec2_{i}_custom", f"s2_multi_spec2_{i}_custom"))
                                else:
                                    st.text_input("规格2内容（手填）", value=_prev_spec, key=f"s2_multi_spec2_{i}", label_visibility="collapsed", on_change=_sync_s2_val, args=(f"s2_multi_spec2_{i}", f"s2_multi_spec2_{i}"))
                    else:
                        _render_s2_field(ec, is_prefilled=True, is_required=_is_req)
        else:
            st.info("🎉 当前模板内暂无需要临时手填的必填项。")

    # ── 图片素材处理区：首图+细节图1-10 一块多图上传按序对应；方块图/色块图/SKU图 固定独立位 ──
    # 批量模式：首图/细节图/方块图/色块图 按「货号」共享；SKU图 按行（卖家SKU）独立
    # 单品模式：所有图片共用一套 key
    _image_cols = [c for c in _excel_cols if c and (
        _tpl.get(c, {}).get("type") == "image" or any(kw in c for kw in IMAGE_COL_KEYWORDS)
    )]
    _main_detail_cols, _other_image_cols = _split_main_detail_image_cols(_image_cols)
    # 批量模式下：_other_image_cols 拆分为「SKU图」（按行独立）和「共享图」（方块图/色块图，按货号共享）
    _sku_img_cols = [c for c in _other_image_cols if c and "SKU图" in c]
    _shared_img_cols = [c for c in _other_image_cols if c not in _sku_img_cols]
    # 兼容单品模式：沿用原 _img_row_idx/_img_key_suffix
    _img_row_idx = (st.session_state.get("s2_form_row_idx", 1) - 1) if _use_batch else None
    _img_key_suffix = f"_row_{_img_row_idx}" if _img_row_idx is not None else ""

    def _save_and_load_img_bytes(uploaded, backup_key):
        """保存上传文件 bytes 到 backup_key；若 uploaded 为空则从 backup 恢复。返回 (uploaded_or_restored, bytes)。"""
        if uploaded:
            try:
                if hasattr(uploaded, "seek"):
                    uploaded.seek(0)
                _b = uploaded.getvalue() if hasattr(uploaded, "getvalue") else (uploaded.read() if hasattr(uploaded, "read") else None)
                _fn = getattr(uploaded, "name", None) or "image.png"
                if _b:
                    st.session_state[backup_key] = (_b, _fn)
                    if hasattr(uploaded, "seek"):
                        uploaded.seek(0)
                return uploaded, _b
            except Exception:
                pass
        bk = st.session_state.get(backup_key)
        if isinstance(bk, tuple) and bk[0]:
            restored = _BytesFile(bk[0], bk[1])
            return restored, bk[0]
        return None, None

    def _preview_file(f, width=80):
        """显示图片预览（兼容 UploadedFile / _BytesFile / tuple(bytes,fn)）。"""
        try:
            if isinstance(f, tuple):
                _b = f[0]
            else:
                if hasattr(f, "seek"):
                    f.seek(0)
                _b = f.getvalue() if hasattr(f, "getvalue") else (f.read() if hasattr(f, "read") else None)
                if _b and hasattr(f, "seek"):
                    f.seek(0)
            if _b:
                st.image(_b, width=width, caption="")
        except Exception:
            pass

    if _image_cols:
        st.markdown("---")
        st.subheader("📸 图片素材处理区 (本地传图转直链)")
        with st.expander("📋 说明与要求", expanded=False):
            st.caption("以下列在模板中设为「本地图片转直链」。侧栏配置 Open Key ID / Secret Key 且 IP 在白名单时，优先走 SHEIN 直传；否则自动用免费图床(sm.ms → Catbox → 0x0.st)生成公网直链并填入 Excel，无需白名单、放线上也可用。SHEIN 要求：主图/细节图 1340×1785 或 1:1 且 900–2200px；方形图 1:1 且 900–2200px；色块图 80×80；JPG/PNG，≤3MB。")

        if _use_batch:
            # ── 批量模式：货号选择器 ──
            _batch_df_for_img = st.session_state.get("s2_edited_batch_df")
            if _batch_df_for_img is None or not isinstance(_batch_df_for_img, pd.DataFrame):
                _batch_df_for_img = st.session_state.get("s2_batch_df")
            _unique_huohao = []
            if _batch_df_for_img is not None and _huohao_col in _batch_df_for_img.columns:
                for _hh in _batch_df_for_img[_huohao_col]:
                    _hh_s = str(_hh or "").strip()
                    if _hh_s and _hh_s not in ("", "None", "nan") and _hh_s not in _unique_huohao:
                        _unique_huohao.append(_hh_s)
            if not _unique_huohao:
                st.info("📋 请先在上方「批量录入表」填写**供方货号**，再来上传图片。")
            else:
                _sel_huohao = st.selectbox(
                    "当前编辑货号（供方货号）——首图/细节图/方块图/色块图按货号共享，SKU图按卖家SKU独立",
                    options=_unique_huohao,
                    key="s2_img_huohao_selector"
                )
                _safe_hh = re.sub(r"[^a-zA-Z0-9_\u4e00-\u9fff]", "_", _sel_huohao)

                # 获取该货号下的所有行 [(df_idx, sku_value)]
                _rows_for_hh = []
                if _batch_df_for_img is not None and _huohao_col in _batch_df_for_img.columns:
                    for _ri, _rrow in _batch_df_for_img.iterrows():
                        if str(_rrow.get(_huohao_col, "") or "").strip() == _sel_huohao:
                            _rsku = str(_rrow.get(_sku_col, "") or "").strip()
                            _rows_for_hh.append((_ri, _rsku))

                # ── 首图 / 细节图（按货号共享）──
                if _main_detail_cols:
                    st.caption(f"**首图 / 细节图**（货号 `{_sel_huohao}` 共用）：第1张→首图，第2张→细节图1，以此类推。")
                    _main_hh_key = f"img_main_detail_list_huohao_{_safe_hh}"
                    _main_backup = f"{_main_hh_key}_backup"
                    _main_list = st.file_uploader(
                        "上传首图及细节图（可多选，按顺序对应）",
                        type=["png", "jpg", "jpeg", "webp"],
                        accept_multiple_files=True,
                        key=_main_hh_key,
                        help="第1张=首图，第2张=细节图1，第3张=细节图2…"
                    )
                    if _main_list:
                        try:
                            _bkp = []
                            for _f in _main_list:
                                if hasattr(_f, "seek"):
                                    _f.seek(0)
                                _b = _f.getvalue() if hasattr(_f, "getvalue") else (_f.read() if hasattr(_f, "read") else None)
                                _fn = getattr(_f, "name", None) or "image.png"
                                if _b:
                                    _bkp.append((_b, _fn))
                            if _bkp:
                                st.session_state[_main_backup] = _bkp
                        except Exception:
                            pass
                    _display_list = _main_list if _main_list else st.session_state.get(_main_backup, [])
                    _n = min(len(_main_detail_cols), 11)
                    if _n > 0:
                        _pcols = st.columns(_n)
                        for _pi in range(_n):
                            with _pcols[_pi]:
                                st.caption(_main_detail_cols[_pi])
                                if _pi < len(_display_list):
                                    _preview_file(_display_list[_pi], width=80)
                                else:
                                    st.caption("—")

                # ── 方块图 / 色块图（按货号共享）──
                if _shared_img_cols:
                    st.caption(f"**方块图 / 色块图**（货号 `{_sel_huohao}` 共用）")
                    _sh_ui_cols = st.columns(len(_shared_img_cols))
                    for _j, _img_col in enumerate(_shared_img_cols):
                        with _sh_ui_cols[_j]:
                            st.caption(f"**{_display_name(_img_col)}**")
                            _sh_key = f"img_{_img_col}_huohao_{_safe_hh}"
                            st.file_uploader("上传", type=["png", "jpg", "jpeg", "webp"],
                                             key=_sh_key, label_visibility="collapsed",
                                             help=f"对应 Excel 列「{_img_col}」（同货号共用）")
                            _sh_up, _sh_b = _save_and_load_img_bytes(
                                st.session_state.get(_sh_key), f"{_sh_key}_backup")
                            if _sh_b:
                                st.image(_sh_b, width=100, caption="预览")

                # ── SKU图（按卖家SKU独立）──
                if _sku_img_cols:
                    st.caption(f"**SKU图**（每个卖家SKU独立上传）— 货号 `{_sel_huohao}` 共 {len(_rows_for_hh)} 个SKU")
                    if not _rows_for_hh:
                        st.info("该货号暂无行数据，请先在批量录入表填写。")
                    for _ri, _rsku in _rows_for_hh:
                        _sku_label = _rsku if _rsku else f"第{_ri+1}行"
                        _sku_c1, _sku_c2 = st.columns([1, len(_sku_img_cols) * 2])
                        with _sku_c1:
                            st.markdown(f"**卖家SKU**  \n`{_sku_label}`")
                        with _sku_c2:
                            _sku_sub = st.columns(len(_sku_img_cols))
                            for _j, _img_col in enumerate(_sku_img_cols):
                                with _sku_sub[_j]:
                                    st.caption(f"**{_display_name(_img_col)}**")
                                    _sk_key = f"img_{_img_col}_row_{_ri}"
                                    st.file_uploader("上传", type=["png", "jpg", "jpeg", "webp"],
                                                     key=_sk_key, label_visibility="collapsed",
                                                     help=f"对应 Excel 列「{_img_col}」（卖家SKU: {_sku_label}）")
                                    _sk_up, _sk_b = _save_and_load_img_bytes(
                                        st.session_state.get(_sk_key), f"{_sk_key}_backup")
                                    if _sk_b:
                                        st.image(_sk_b, width=100, caption="预览")

        else:
            # ── 单品模式 ──
            # 首图 / 细节图：始终共用（复色多变体也共用同一套首图/细节图）
            if _main_detail_cols:
                st.caption("**首图 / 细节图**：一次可多选，第 1 张→首图，第 2 张→细节图1，第 3 张→细节图2… 以此类推。")
                _main_list = st.file_uploader(
                    "上传首图及细节图（可多选，按顺序对应）",
                    type=["png", "jpg", "jpeg", "webp"],
                    accept_multiple_files=True,
                    key="img_main_detail_list",
                    help="第1张=首图，第2张=细节图1，第3张=细节图2…"
                )
                _list = _main_list if _main_list is not None else []
                _n = len(_main_detail_cols)
                if _n > 0:
                    _preview_cols = st.columns(min(_n, 11))
                    for _i in range(min(_n, 11)):
                        with _preview_cols[_i]:
                            st.caption(_main_detail_cols[_i])
                            if _i < len(_list):
                                _preview_file(_list[_i], width=80)
                            else:
                                st.caption("—")

            if _is_multi_variant and _sku_img_cols:
                # ── 复色单条：方块图/色块图 所有SKU共用；SKU图 每个SKU独立 ──
                if _shared_img_cols:
                    st.caption("**方块图 / 色块图**（所有SKU共用）")
                    _sh_fix_cols = st.columns(len(_shared_img_cols))
                    for _j, _img_col in enumerate(_shared_img_cols):
                        with _sh_fix_cols[_j]:
                            st.caption(f"**{_display_name(_img_col)}**")
                            st.file_uploader(
                                "上传", type=["png", "jpg", "jpeg", "webp"],
                                key=f"img_{_img_col}",
                                help=f"对应 Excel 列「{_img_col}」（所有SKU共用）",
                                label_visibility="collapsed"
                            )
                            _sh_up, _sh_b = _save_and_load_img_bytes(
                                st.session_state.get(f"img_{_img_col}"), f"img_{_img_col}_backup")
                            if _sh_b:
                                st.image(_sh_b, width=100, caption="预览")

                st.caption(f"**SKU图**（每个卖家SKU独立上传）— 共 {len(_variant_skus)} 个SKU")
                for _vi, _vsku in enumerate(_variant_skus):
                    _vsku_label = _vsku if _vsku else f"SKU #{_vi + 1}"
                    _vc1, _vc2 = st.columns([1, len(_sku_img_cols) * 2])
                    with _vc1:
                        st.markdown(f"**卖家SKU**  \n`{_vsku_label}`")
                    with _vc2:
                        _sku_sub_cols = st.columns(len(_sku_img_cols))
                        for _j, _img_col in enumerate(_sku_img_cols):
                            with _sku_sub_cols[_j]:
                                st.caption(f"**{_display_name(_img_col)}**")
                                _sk_key = f"img_{_img_col}_single_sku_{_vi}"
                                st.file_uploader(
                                    "上传", type=["png", "jpg", "jpeg", "webp"],
                                    key=_sk_key,
                                    help=f"对应 Excel 列「{_img_col}」（卖家SKU: {_vsku_label}）",
                                    label_visibility="collapsed"
                                )
                                _sk_up, _sk_b = _save_and_load_img_bytes(
                                    st.session_state.get(_sk_key), f"{_sk_key}_backup")
                                if _sk_b:
                                    st.image(_sk_b, width=100, caption="预览")
            else:
                # ── 单色单品 或 无SKU图：方块图/色块图/SKU图 全部共用一套 key ──
                if _other_image_cols:
                    st.caption("**方块图 / 色块图 / SKU图**：各有一个固定上传位。")
                    _fix_cols = st.columns(len(_other_image_cols))
                    for _j, _img_col in enumerate(_other_image_cols):
                        with _fix_cols[_j]:
                            st.caption(f"**{_display_name(_img_col)}**")
                            st.file_uploader(
                                "上传",
                                type=["png", "jpg", "jpeg", "webp"],
                                key=f"img_{_img_col}",
                                help=f"对应 Excel 列「{_img_col}」",
                                label_visibility="collapsed"
                            )
                            _up_single, _b_single = _save_and_load_img_bytes(
                                st.session_state.get(f"img_{_img_col}"), f"img_{_img_col}_backup")
                            if _b_single:
                                st.image(_b_single, width=100, caption="预览")

    # 单条生成时才展示模板已固定区；批量生成时以表格/表单式录入为准
    if not _use_batch:
        st.markdown("---")
        st.subheader(f"📝 4. 模板已固定的规格与属性 ({len(_prefilled_fixed_cols)} 项)")
        st.caption("以下为第一步中已配置的其它固定信息（规格1/规格1内容等已放在上方「核心动态录入区」）。可直接在此核对与修改，生成 Excel 时以此处最新值为准。")
        if _prefilled_fixed_cols:
            with st.expander("📌 展开核对或修改", expanded=True):
                for ec in _prefilled_fixed_cols:
                    _render_s2_field(ec, is_prefilled=True)
        else:
            st.info("📂 当前模板没有任何固定的规格或属性。")

    # ── 5. 生成按钮 ──────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📥 5. 拉取飞书数据并生成 Excel")

    # 检查是否有原始 Excel 模板（第一步上传的，或加载模板时自动恢复的）
    _has_excel = bool(st.session_state.get('uploaded_excel_bytes'))

    if not _has_excel:
        st.warning("⚠️ 未检测到 Excel 模板。请先在「1. 选择模板」中加载一个曾在第一步保存过 Excel 的配置模板，或返回第一步上传 Excel 并保存模板。")

    if st.button("🚀 拉取飞书并生成 Excel", type="primary", use_container_width=True,
                 disabled=not _has_excel):
        # 清除上一次生成结果，避免重新生成前显示旧的下载按钮
        st.session_state.pop("s2_excel_ready", None)
        st.session_state.pop("s2_excel_filename", None)
        _batch_row = None
        if _use_batch:
            # 表格与表单已联动，统一使用同一份 DataFrame 数据生成
            _edited_df = st.session_state.get("s2_edited_batch_df")
            if _edited_df is None or _edited_df.empty:
                st.error("❌ 批量表为空，请至少填写一行（供方货号、卖家SKU、标题、五点描述等）。")
                st.stop()
            _rows_to_process = _edited_df[_edited_df[_sku_col].astype(str).str.strip() != ""].copy()
            if _rows_to_process.empty:
                st.error("❌ 请至少在一行中填写卖家SKU。")
                st.stop()
            _all_rows_to_write = []
            _all_merge_log = []
        else:
            if not _input_sku:
                st.error("❌ 请先输入卖家SKU作为检索凭证！")
                st.stop()
        if not _use_batch and _is_multi_variant:
            _extra_cnt = st.session_state.get("s2_extra_sku_count", 0)
            _n_check = 1 + _extra_cnt
            _skus_check = [(_input_sku or "").strip()] + [(st.session_state.get(f"s2_extra_sku_{i}", "") or "").strip() for i in range(_extra_cnt)]
            def _effective_spec2(i):
                _c = (st.session_state.get(f"s2_multi_spec2_{i}_custom", "") or "").strip()
                _s = st.session_state.get(f"s2_multi_spec2_{i}", "")
                if _c:
                    return _c
                if _s and _s not in ("-- 请选择 --", "-- 不填 --", "-- 自己填写 --"):
                    return _s
                return ""
            _spec2_check = [_effective_spec2(i) for i in range(_n_check)]
            if _n_check < 2 or not _input_huohao:
                st.error("❌ 复色模式下请填写供方货号（所有行共用），并至少点击一次「添加复色卖家SKU」填写至少一个复色 SKU。")
                st.stop()
            if not all(_skus_check):
                st.error("❌ 复色模式下请填写每一个变体的卖家SKU（主 SKU + 上方添加的复色 SKU 均需填写）。")
                st.stop()
            if not all(_spec2_check):
                st.error("❌ 复色模式下请为每个变体填写规格2内容。")
                st.stop()

        with st.spinner("正在从飞书拉取数据..."):
            try:
                # Step 1: 获取 Token
                _token = get_feishu_tenant_token(_fs_auth["app_id"], _fs_auth["app_secret"])

                import openpyxl
                from io import BytesIO
                _original_bytes = st.session_state.get('uploaded_excel_bytes')
                if not _original_bytes:
                    st.error("原始 Excel 文件数据丢失，请返回第一步重新上传。")
                    st.stop()

                _image_urls = {}
                if _use_batch:
                    # 批量模式：表格与表单已联动，统一按 DataFrame 行逐条拉飞书、合并，写入同一 Excel
                    wb = openpyxl.load_workbook(BytesIO(_original_bytes))
                    ws = wb.worksheets[1]
                    _excel_cols = _excel_header_row(ws)
                    _image_cols_gen = [c for c in _excel_cols if c and (
                        _tpl.get(c, {}).get("type") == "image" or any(kw in c for kw in IMAGE_COL_KEYWORDS)
                    )]
                    _main_detail_cols_gen, _ = _split_main_detail_image_cols(_image_cols_gen)
                    _all_rows_to_write = []
                    _all_merge_log = []

                    # 批量预拉飞书：并行请求所有行的飞书记录，避免逐行串行等待
                    def _fetch_one_feishu(args):
                        _ri, _sku = args
                        _s = (_sku or "").strip()
                        if not _s:
                            return (_ri, None)
                        try:
                            _rec = search_feishu_record(
                                _token, _fs_auth["app_token"], _fs_auth["table_id"],
                                _match_field, _s
                            )
                            return (_ri, _rec)
                        except Exception:
                            return (_ri, None)
                    _feishu_tasks = [(_idx, str(_row.get(_sku_col, "") or "").strip()) for _idx, _row in _rows_to_process.iterrows()]
                    _feishu_by_idx = {}
                    with ThreadPoolExecutor(max_workers=min(8, max(1, len(_feishu_tasks)))) as _ex:
                        for _ri, _rec in _ex.map(_fetch_one_feishu, _feishu_tasks):
                            _feishu_by_idx[_ri] = _rec

                    # 批量：每行复用与「单品上传」完全相同的图片转链逻辑。
                    # 首图/细节图/方块图/色块图 按货号共享，SKU图 按行独立
                    _gen_sku_img_cols = [c for c in _image_cols_gen if c and "SKU图" in c]
                    _gen_shared_img_cols = [c for c in _image_cols_gen if c not in _gen_sku_img_cols]

                    def _get_shared_uploaded(cn, safe_hh):
                        """按货号读取共享图（首图/细节图/方块图/色块图）。"""
                        if cn in _main_detail_cols_gen:
                            _hh_key = f"img_main_detail_list_huohao_{safe_hh}"
                            _md_list = st.session_state.get(_hh_key, []) or []
                            if not _md_list:
                                _md_list = st.session_state.get(f"{_hh_key}_backup", [])
                            _md_i = _main_detail_cols_gen.index(cn)
                            if _md_i >= len(_md_list):
                                return None
                            _item = _md_list[_md_i]
                            if isinstance(_item, tuple):
                                return _BytesFile(_item[0], _item[1]) if _item[0] else None
                            return _item
                        # 方块图/色块图
                        _sh_key = f"img_{cn}_huohao_{safe_hh}"
                        _up = st.session_state.get(_sh_key)
                        if _up is None:
                            _bk = st.session_state.get(f"{_sh_key}_backup")
                            if isinstance(_bk, tuple) and _bk[0]:
                                return _BytesFile(_bk[0], _bk[1])
                        return _up

                    def _get_sku_uploaded(cn, row_idx):
                        """按行读取 SKU 图。"""
                        _up = st.session_state.get(f"img_{cn}_row_{row_idx}")
                        if _up is None:
                            _bk = st.session_state.get(f"img_{cn}_row_{row_idx}_backup")
                            if isinstance(_bk, tuple) and _bk[0]:
                                return _BytesFile(_bk[0], _bk[1])
                        return _up

                    for _pos, (_idx, _row) in enumerate(_rows_to_process.iterrows()):
                        _row_image_urls = {}
                        _row_huohao_raw = str(_row.get(_huohao_col, "") or "").strip()
                        _row_safe_hh = re.sub(r"[^a-zA-Z0-9_\u4e00-\u9fff]", "_", _row_huohao_raw)
                        if not USE_SHEIN_IMAGE_API and _image_cols_gen:
                            def _get_uploaded_for_this_row(cn, _shh=_row_safe_hh, _ridx=_idx):
                                if cn in _gen_sku_img_cols:
                                    return _get_sku_uploaded(cn, _ridx)
                                return _get_shared_uploaded(cn, _shh)
                            _row_image_urls = _upload_image_cols_to_urls_parallel(_image_cols_gen, lambda c: _get_uploaded_for_this_row(c))
                        _input_huohao = str(_row.get(_huohao_col, "") or "").strip()
                        _input_sku = str(_row.get(_sku_col, "") or "").strip()
                        _batch_row = _row.to_dict()
                        def _batch_cell(cn):
                            if cn in _spec_content_cols:
                                v = str(_batch_row.get(cn + "_手填") or "").strip() or str(_batch_row.get(cn) or "").strip()
                            else:
                                v = str(_batch_row.get(cn) or "").strip()
                            if v in ("-- 请选择 --", "-- 不填 --", "-- 自己填写 --"):
                                return ""
                            return v
                        _record = _feishu_by_idx.get(_idx)
                        if _record is None:
                            _all_merge_log.append(f"[行{_pos+1}] 未找到飞书: {_input_sku}")
                            continue
                        _merge_log = []
                        row_1_data = {}
                        for col_idx, col_name in enumerate(_excel_cols):
                            if not col_name:
                                row_1_data[col_name] = ""
                                continue
                            cfg = _tpl.get(col_name, {})
                            cfg_type = cfg.get("type", "ignore")
                            final_value = ""
                            if cfg_type in ("fixed_single", "fixed_multi", "fixed_text"):
                                if _batch_row is not None and col_name in _batch_df_columns:
                                    final_value = _batch_cell(col_name) or _s2_manual_values.get(col_name, "")
                                elif col_name in _s2_manual_values:
                                    final_value = _s2_manual_values[col_name]
                                else:
                                    template_val = cfg.get("value", "")
                                    if template_val in ("-- 请选择 --", "-- 不填 --"):
                                        template_val = ""
                                    final_value = template_val
                                _merge_log.append(f"✅ {col_name}: 固定值 → \'{final_value}\'")
                            elif cfg_type == "feishu":
                                feishu_key = cfg.get("feishu_key", "")
                                if feishu_key and feishu_key in _record:
                                    final_value = _record[feishu_key]
                                    _merge_log.append(f"🔗 {col_name}: 飞书[{feishu_key}] → \'{final_value}\'")
                                else:
                                    _merge_log.append(f"⚠️ {col_name}: 飞书字段[{feishu_key}]未找到")
                            elif cfg_type == "price_x1.5":
                                # 取价格列的值 * 1.5 并取整
                                _price_col_ref = next((c for c in _excel_cols if c and "价格" in c and "建议" not in c), None)
                                _price_raw = ""
                                if _price_col_ref:
                                    _price_raw = str(row_1_data.get(_price_col_ref, "") or "").strip()
                                    if not _price_raw:
                                        _p_cfg = _tpl.get(_price_col_ref, {})
                                        if _p_cfg.get("type") == "feishu":
                                            _p_fk = _p_cfg.get("feishu_key", "")
                                            if _p_fk and _p_fk in _record:
                                                _price_raw = str(_record[_p_fk]).strip()
                                if _price_raw:
                                    try:
                                        final_value = str(int(round(float(_price_raw.replace(",", "")) * 1.5)))
                                        _merge_log.append(f"💹 {col_name}: 价格×1.5 ({_price_raw}×1.5) → '{final_value}'")
                                    except (ValueError, TypeError):
                                        _merge_log.append(f"⚠️ {col_name}: 价格×1.5 计算失败，原价格='{_price_raw}'")
                                else:
                                    _merge_log.append(f"⚠️ {col_name}: 未找到价格列值，跳过×1.5")
                            elif cfg_type == "manual":
                                if _batch_row is not None and col_name in _batch_df_columns:
                                    final_value = _batch_cell(col_name) or _s2_manual_values.get(col_name, "")
                                elif "卖家SKU" in col_name or "SKU" in col_name.upper():
                                    final_value = _input_sku
                                elif "货号" in col_name:
                                    final_value = _input_huohao
                                else:
                                    final_value = _s2_manual_values.get(col_name, "")
                                _merge_log.append(f"✏️ {col_name}: 手填 → \'{final_value}\'")
                            elif cfg_type == "image" or col_name in _image_cols_gen:
                                # SKU图按行读；首图/细节图/方块图/色块图按货号读
                                if col_name in _gen_sku_img_cols:
                                    _uploaded = _get_sku_uploaded(col_name, _idx)
                                elif col_name in _main_detail_cols_gen:
                                    _hh_key_gen = f"img_main_detail_list_huohao_{_row_safe_hh}"
                                    _md_list = st.session_state.get(_hh_key_gen, []) or []
                                    if not _md_list:
                                        _md_list = st.session_state.get(f"{_hh_key_gen}_backup", [])
                                    _md_idx = _main_detail_cols_gen.index(col_name)
                                    _uploaded = None
                                    if _md_idx < len(_md_list):
                                        _item = _md_list[_md_idx]
                                        _uploaded = _BytesFile(_item[0], _item[1]) if isinstance(_item, tuple) and _item[0] else _item
                                else:
                                    _uploaded = _get_shared_uploaded(col_name, _row_safe_hh)
                                _shein_auth = _get_shein_from_secrets() or st.session_state.get("shein_auth") or {}
                                _itype = _tpl.get(col_name, {}).get("shein_image_type") or _shein_image_type_for_column(col_name)
                                if USE_SHEIN_IMAGE_API and _uploaded and _shein_auth.get("open_key_id") and _shein_auth.get("secret_key"):
                                    _file_bytes = _uploaded.getvalue() if hasattr(_uploaded, "getvalue") else (_uploaded.read() if hasattr(_uploaded, "read") else None)
                                    _fname = getattr(_uploaded, "name", None) or "image.png"
                                    _up_err = None
                                    if _file_bytes:
                                        final_value, _up_err = shein_upload_pic(
                                            _shein_auth["open_key_id"], _shein_auth["secret_key"],
                                            _file_bytes, _fname, _itype
                                        )
                                    else:
                                        final_value = ""
                                    if not final_value:
                                        _public_url = upload_image_to_url(None, _file_bytes, _fname)
                                        _tv, _tf_err = shein_transform_pic(
                                            _shein_auth["open_key_id"], _shein_auth["secret_key"], _public_url, _itype
                                        )
                                        final_value = _tv or _public_url
                                        if not final_value and _up_err:
                                            _merge_log.append(f"🖼️ {col_name}: 直传失败 {_up_err}")
                                        if not final_value and _tf_err:
                                            _merge_log.append(f"🖼️ {col_name}: 转链失败 {_tf_err}")
                                    _url_display = (final_value[:60] + "...") if len(final_value) > 60 else final_value
                                    _merge_log.append(f"🖼️ {col_name}: 本地图→SHEIN直链 → '{_url_display}'")
                                else:
                                    if _row_image_urls:
                                        final_value = _row_image_urls.get(col_name, "")
                                    else:
                                        if _uploaded and hasattr(_uploaded, "seek"):
                                            _uploaded.seek(0)
                                        final_value = upload_image_to_url(_uploaded) if _uploaded else ""
                                    _had_file = (col_name in _row_image_urls) or (_uploaded is not None)
                                    if final_value:
                                        _merge_log.append(f"🖼️ {col_name}: 本地图→直链 → '{final_value[:60] + '...' if len(final_value) > 60 else final_value}'")
                                    else:
                                        _merge_log.append(f"🖼️ {col_name}: 本地图→直链 → **{'上传失败' if _had_file else '未上传'}**")
                            else:
                                _merge_log.append(f"⬜ {col_name}: 忽略")
                            row_1_data[col_name] = final_value
                        _all_rows_to_write.append(row_1_data)
                        _all_merge_log.extend(_merge_log)
                    # 批量图片说明：已按与单品相同逻辑逐行生成链接，详见下方每行的 🖼️ 日志
                    _image_summary = ["--- [批量图片] 按与「单品上传」相同逻辑逐行生成链接 ---"] if _image_cols_gen else []
                    rows_to_write = _all_rows_to_write
                    # 调试：确认每行的图片列是否真的写进了 row_1_data
                    _debug_summary = [f"[调试] 批量共 {len(rows_to_write)} 行，图片列={_image_cols_gen}"]
                    for _di, _rd in enumerate(rows_to_write):
                        for _ic in _image_cols_gen:
                            _dv = _rd.get(_ic, "__未找到key__")
                            _debug_summary.append(f"[调试] 第{_di+1}行 {_ic} = '{_dv}'")
                    _merge_log = (_debug_summary + (_image_summary + _all_merge_log if _image_summary else _all_merge_log))
                else:
                    # 单条模式：按原逻辑拉一条记录、合并、复色裂变
                    try:
                        _record = search_feishu_record(
                            _token,
                            _fs_auth["app_token"],
                            _fs_auth["table_id"],
                            _match_field,
                            (_input_sku or "").strip()
                        )
                    except ValueError as ve:
                        st.error(f"❌ {ve}")
                        st.stop()
                    if _record is None:
                        st.error(f"❌ 未在飞书中找到匹配「{_match_field} = {_input_sku}」的数据！请检查卖家SKU是否正确。")
                        st.stop()
                    st.success(f"✅ 成功找到匹配记录！飞书返回了 {len(_record)} 个字段。")
                    with st.expander("🔍 飞书返回数据预览", expanded=False):
                        st.json(_record)

                    wb = openpyxl.load_workbook(BytesIO(_original_bytes))
                    ws = wb.worksheets[1]
                    _excel_cols = _excel_header_row(ws)
                    _image_cols_gen = [c for c in _excel_cols if c and (
                        _tpl.get(c, {}).get("type") == "image" or any(kw in c for kw in IMAGE_COL_KEYWORDS)
                    )]
                    _main_detail_cols_gen, _ = _split_main_detail_image_cols(_image_cols_gen)

                    _image_urls = {}
                    if not USE_SHEIN_IMAGE_API and _image_cols_gen:
                        def _get_uploaded_for_col(cn):
                            if cn in _main_detail_cols_gen:
                                _md_list = st.session_state.get("img_main_detail_list", [])
                                _idx = _main_detail_cols_gen.index(cn)
                                return _md_list[_idx] if _idx < len(_md_list) else None
                            # 复色单条模式：SKU图第0个变体用独立 key
                            if _is_multi_variant and cn in _sku_img_cols:
                                _sk0 = f"img_{cn}_single_sku_0"
                                _up = st.session_state.get(_sk0)
                                if _up is None:
                                    _bk = st.session_state.get(f"{_sk0}_backup")
                                    if isinstance(_bk, tuple) and _bk[0]:
                                        return _BytesFile(_bk[0], _bk[1])
                                return _up
                            return _get_uploaded_for_image_col(cn)
                        _image_urls = _upload_image_cols_to_urls_parallel(_image_cols_gen, _get_uploaded_for_col)

                    _merge_log = []
                    row_1_data = {}
                if not _use_batch:
                    for col_idx, col_name in enumerate(_excel_cols):
                        if not col_name:
                            row_1_data[col_name] = ""
                            continue
                        cfg = _tpl.get(col_name, {})
                        cfg_type = cfg.get("type", "ignore")
                        final_value = ""
                        if cfg_type in ("fixed_single", "fixed_multi", "fixed_text"):
                            if _batch_row is not None and col_name in _batch_df_columns:
                                final_value = (_batch_row.get(col_name) or "") or _s2_manual_values.get(col_name, "")
                            elif col_name in _s2_manual_values:
                                final_value = _s2_manual_values[col_name]
                            else:
                                template_val = cfg.get("value", "")
                                if template_val in ("-- 请选择 --", "-- 不填 --"):
                                    template_val = ""
                                final_value = template_val
                            _merge_log.append(f"✅ {col_name}: 固定值 → \'{final_value}\'")
                        elif cfg_type == "feishu":
                            feishu_key = cfg.get("feishu_key", "")
                            if feishu_key and feishu_key in _record:
                                final_value = _record[feishu_key]
                                _merge_log.append(f"🔗 {col_name}: 飞书[{feishu_key}] → \'{final_value}\'")
                            else:
                                _merge_log.append(f"⚠️ {col_name}: 飞书字段[{feishu_key}]未找到")
                        elif cfg_type == "price_x1.5":
                            # 取价格列的值 * 1.5 并取整
                            _price_col_ref = next((c for c in _excel_cols if c and "价格" in c and "建议" not in c), None)
                            _price_raw = ""
                            if _price_col_ref:
                                _price_raw = str(row_1_data.get(_price_col_ref, "") or "").strip()
                                if not _price_raw:
                                    _p_cfg = _tpl.get(_price_col_ref, {})
                                    if _p_cfg.get("type") == "feishu":
                                        _p_fk = _p_cfg.get("feishu_key", "")
                                        if _p_fk and _p_fk in _record:
                                            _price_raw = str(_record[_p_fk]).strip()
                            if _price_raw:
                                try:
                                    final_value = str(int(round(float(_price_raw.replace(",", "")) * 1.5)))
                                    _merge_log.append(f"💹 {col_name}: 价格×1.5 ({_price_raw}×1.5) → '{final_value}'")
                                except (ValueError, TypeError):
                                    _merge_log.append(f"⚠️ {col_name}: 价格×1.5 计算失败，原价格='{_price_raw}'")
                            else:
                                _merge_log.append(f"⚠️ {col_name}: 未找到价格列值，跳过×1.5")
                        elif cfg_type == "manual":
                            if _batch_row is not None and col_name in _batch_df_columns:
                                final_value = (_batch_row.get(col_name) or "") or _s2_manual_values.get(col_name, "")
                            elif "卖家SKU" in col_name or "SKU" in col_name.upper():
                                final_value = _input_sku
                            elif "货号" in col_name:
                                final_value = _input_huohao
                            else:
                                final_value = _s2_manual_values.get(col_name, "")
                            _merge_log.append(f"✏️ {col_name}: 手填 → \'{final_value}\'")
                        elif cfg_type == "image" or col_name in _image_cols_gen:
                            if col_name in _main_detail_cols_gen:
                                _md_list = st.session_state.get("img_main_detail_list", [])
                                _md_idx = _main_detail_cols_gen.index(col_name)
                                _uploaded = _md_list[_md_idx] if _md_idx < len(_md_list) else None
                            elif _is_multi_variant and col_name in _sku_img_cols:
                                # 复色单条模式：SKU图第0个变体用独立 key
                                _sk0 = f"img_{col_name}_single_sku_0"
                                _up = st.session_state.get(_sk0)
                                if _up is None:
                                    _bk = st.session_state.get(f"{_sk0}_backup")
                                    if isinstance(_bk, tuple) and _bk[0]:
                                        _up = _BytesFile(_bk[0], _bk[1])
                                _uploaded = _up
                            else:
                                _uploaded = _get_uploaded_for_image_col(col_name)
                            _shein_auth = _get_shein_from_secrets() or st.session_state.get("shein_auth") or {}
                            _itype = _tpl.get(col_name, {}).get("shein_image_type")
                            if _itype is None:
                                _itype = _shein_image_type_for_column(col_name)
                            if USE_SHEIN_IMAGE_API and _uploaded and _shein_auth.get("open_key_id") and _shein_auth.get("secret_key"):
                                _file_bytes = _uploaded.getvalue() if hasattr(_uploaded, "getvalue") else (_uploaded.read() if hasattr(_uploaded, "read") else None)
                                _fname = getattr(_uploaded, "name", None) or "image.png"
                                _up_err = None
                                if _file_bytes:
                                    final_value, _up_err = shein_upload_pic(
                                        _shein_auth["open_key_id"], _shein_auth["secret_key"],
                                        _file_bytes, _fname, _itype
                                    )
                                else:
                                    final_value = ""
                                if not final_value:
                                    _public_url = upload_image_to_url(None, _file_bytes, _fname)
                                    _tv, _tf_err = shein_transform_pic(
                                        _shein_auth["open_key_id"], _shein_auth["secret_key"], _public_url, _itype
                                    )
                                    final_value = _tv or _public_url
                                    if not final_value and _up_err:
                                        _merge_log.append(f"🖼️ {col_name}: 直传失败 {_up_err}")
                                    if not final_value and _tf_err:
                                        _merge_log.append(f"🖼️ {col_name}: 转链失败 {_tf_err}")
                                _url_display = (final_value[:60] + "...") if len(final_value) > 60 else final_value
                                _merge_log.append(f"🖼️ {col_name}: 本地图→SHEIN直链(type={_itype}) → '{_url_display}'")
                            else:
                                if _image_urls:
                                    final_value = _image_urls.get(col_name, "")
                                else:
                                    if _uploaded and hasattr(_uploaded, "seek"):
                                        _uploaded.seek(0)
                                    final_value = upload_image_to_url(_uploaded) if _uploaded else ""
                                _had_file = (col_name in _image_urls) or (_uploaded is not None)
                                if final_value:
                                    _merge_log.append(f"🖼️ {col_name}: 本地图→直链 → '{final_value[:60] + '...' if len(final_value) > 60 else final_value}'")
                                else:
                                    _merge_log.append(f"🖼️ {col_name}: 本地图→直链 → **{'上传失败' if _had_file else '未上传'}**")
                        else:
                            _merge_log.append(f"⬜ {col_name}: 忽略")
                        row_1_data[col_name] = final_value

                # 仅单条模式：复色逻辑会覆盖 rows_to_write；批量模式已在上方设好 rows_to_write = _all_rows_to_write，此处不再执行
                if not _use_batch:
                    # 复色模式：N 行，供方货号全部相同，每行卖家SKU + 规格2内容不同，飞书字段按每行 SKU 拉取
                    _huohao_col = next((c for c in _excel_cols if c and ("供方货号" in c or "货号" in c)), None)
                    _sku_col = next((c for c in _excel_cols if c and ("卖家SKU" in c or "SKU" in c.upper())), None)
                    _spec2_col = next((c for c in _excel_cols if c and "规格2内容" in c), None)

                    rows_to_write = []
                    if _is_multi_variant:
                        _extra_final = st.session_state.get("s2_extra_sku_count", 0)
                        _n_final = 1 + _extra_final
                        _skus_final = [_input_sku] + [st.session_state.get(f"s2_extra_sku_{i}", "") for i in range(_extra_final)]
                        def _effective_spec2_final(i):
                            _c = (st.session_state.get(f"s2_multi_spec2_{i}_custom", "") or "").strip()
                            _s = st.session_state.get(f"s2_multi_spec2_{i}", "")
                            if _c:
                                return _c
                            if _s and _s not in ("-- 请选择 --", "-- 不填 --", "-- 自己填写 --"):
                                return _s
                            return ""
                        _spec2_final = [_effective_spec2_final(i) for i in range(_n_final)]
                        _merge_log.append(f"[复色] 共 {_n_final} 行，供方货号统一为: {_input_huohao}")
                        for i in range(_n_final):
                            if i == 0:
                                _row = dict(row_1_data)
                                if _huohao_col:
                                    _row[_huohao_col] = _input_huohao
                                if _sku_col:
                                    _row[_sku_col] = _skus_final[0]
                                if _spec2_col and _n_final > 0:
                                    _row[_spec2_col] = _spec2_final[0]
                                rows_to_write.append(_row)
                                _merge_log.append(f"[复色] 第1行: 供方货号={_input_huohao}, 卖家SKU={_skus_final[0]}, 规格2内容={_spec2_final[0] if _spec2_final else ''}")
                            else:
                                # 复色第 2+ 行：深拷贝第 1 行（含已填写的本地图片直链），再覆盖货号/SKU/规格2及飞书字段
                                _row = copy.deepcopy(row_1_data)
                                if _huohao_col:
                                    _row[_huohao_col] = _input_huohao
                                if _sku_col:
                                    _row[_sku_col] = _skus_final[i]
                                if _spec2_col and i < len(_spec2_final):
                                    _row[_spec2_col] = _spec2_final[i]
                                try:
                                    _rec_i = search_feishu_record(_token, _fs_auth["app_token"], _fs_auth["table_id"], _match_field, (_skus_final[i] or "").strip())
                                except ValueError:
                                    _rec_i = None
                                if _rec_i:
                                    for _col_name in _excel_cols:
                                        if not _col_name:
                                            continue
                                        _cfg = _tpl.get(_col_name, {})
                                        if _cfg.get("type") == "feishu":
                                            _fk = _cfg.get("feishu_key", "")
                                            if _fk and _fk in _rec_i:
                                                _row[_col_name] = _rec_i[_fk]
                                # 复色单条：覆盖 SKU图（每个变体独立上传，第 i 个变体用 _single_sku_{i}）
                                for _img_col in _sku_img_cols:
                                    _sk_key_i = f"img_{_img_col}_single_sku_{i}"
                                    _up_ski = st.session_state.get(_sk_key_i)
                                    if _up_ski is None:
                                        _bk_ski = st.session_state.get(f"{_sk_key_i}_backup")
                                        if isinstance(_bk_ski, tuple) and _bk_ski[0]:
                                            _up_ski = _BytesFile(_bk_ski[0], _bk_ski[1])
                                    if _up_ski:
                                        if hasattr(_up_ski, "seek"):
                                            _up_ski.seek(0)
                                        _sku_url_i = upload_image_to_url(_up_ski)
                                        if _sku_url_i:
                                            _row[_img_col] = _sku_url_i
                                            _merge_log.append(f"🖼️ {_img_col}: SKU图(变体{i+1}) → '{_sku_url_i[:60]}'")
                                rows_to_write.append(_row)
                                _merge_log.append(f"[复色] 第{i+1}行: 供方货号={_input_huohao}, 卖家SKU={_skus_final[i]}, 规格2内容={_spec2_final[i] if i < len(_spec2_final) else ''}")
                    else:
                        if _huohao_col:
                            row_1_data[_huohao_col] = _input_huohao
                        rows_to_write = [row_1_data]

                # 写入前取消数据区（第7行起）所有合并单元格，防止第二行及以后被合并遮挡
                _data_start_row = 7
                _merged = getattr(ws, "merged_cells", None)
                if _merged is not None and hasattr(_merged, "ranges"):
                    for _mr in list(_merged.ranges):
                        try:
                            _min_col, _min_row, _max_col, _max_row = _mr.bounds
                            if _max_row >= _data_start_row:
                                ws.unmerge_cells(str(_mr))
                        except Exception:
                            pass
                # 数值列关键词：价格→保留2位小数；库存/重量→取整
                _PRICE_KW  = ("price", "cost", "价格", "售价", "单价")
                _INT_KW    = ("stock", "库存", "weight", "重量")

                def _fmt_numeric_val(col, raw):
                    """对特定列做数值格式化：价格保留2位小数，库存/重量取整，其余原样。"""
                    if not raw and raw != 0:
                        return raw
                    s = str(raw).strip()
                    if not s or s.lower() in ("none", "nan", ""):
                        return ""
                    col_l = (col or "").lower()
                    is_price = any(k in col_l for k in _PRICE_KW)
                    is_int   = any(k in col_l for k in _INT_KW)
                    if is_price or is_int:
                        try:
                            num = float(s)
                            if is_int:
                                return str(int(round(num)))
                            else:  # price
                                return f"{num:.2f}"
                        except (ValueError, TypeError):
                            pass
                    return s

                # 按列顺序写入多行（第 7 行起），统一转为字符串避免 None 或类型导致不显示
                _write_debug = [f"[写入调试] ws={ws.title}, 共{len(rows_to_write)}行"]
                for row_index, row_dict in enumerate(rows_to_write):
                    excel_row = _data_start_row + row_index
                    _img_written = []
                    _img_missing = []
                    for col_idx, col_name in enumerate(_excel_cols):
                        val = row_dict.get(col_name) or row_dict.get((col_name or "").strip())
                        if val is None:
                            val = ""
                        else:
                            val = _fmt_numeric_val(col_name, val)
                            if not isinstance(val, str):
                                val = str(val)
                        ws.cell(row=excel_row, column=col_idx + 1, value=val)
                        if col_name in (_image_cols_gen if _use_batch else []):
                            if val:
                                _img_written.append(f"{col_name}={val[:30]}")
                            else:
                                _img_missing.append(col_name)
                    _write_debug.append(f"[写入调试] 第{row_index+1}行→Excel第{excel_row}行")
                    if _img_written:
                        _write_debug.append(f"  ✅ 有链接: {_img_written}")
                    if _img_missing:
                        _write_debug.append(f"  ❌ 无链接: {_img_missing}")

                # 导出为 BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # 数据合并日志
                with st.expander("📋 数据合并详情", expanded=False):
                    for log_entry in (_write_debug + list(_merge_log)):
                        st.text(log_entry)

                _download_name = (f"SHEIN_上传_批量_{len(rows_to_write)}条.xlsx" if _use_batch
                                 else f"SHEIN_上传_{_input_sku}" + ("_复色" if _is_multi_variant else "") + ".xlsx")
                # 将生成结果存入 session_state，下载按钮在 if 块外持久显示
                st.session_state["s2_excel_ready"] = output.getvalue()
                st.session_state["s2_excel_filename"] = _download_name

                st.balloons()

            except requests.exceptions.HTTPError as he:
                st.error(f"网络异常 (HTTP {he.response.status_code}): {str(he)}")
            except Exception as e:
                st.error(f"❌ 生成失败: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

    # 下载按钮：只要 session_state 有生成结果就持久显示，点击后自动清除
    if st.session_state.get("s2_excel_ready"):
        def _clear_excel_ready():
            st.session_state.pop("s2_excel_ready", None)
            st.session_state.pop("s2_excel_filename", None)
        st.download_button(
            label="📥 下载生成的 Excel 文件",
            data=st.session_state["s2_excel_ready"],
            file_name=st.session_state.get("s2_excel_filename", "output.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            type="primary",
            use_container_width=True,
            on_click=_clear_excel_ready
        )

